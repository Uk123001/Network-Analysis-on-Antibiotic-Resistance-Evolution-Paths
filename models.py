"""
Step 5: Model Evaluation with Comprehensive LIME & SHAP Analysis
Reads Data_Split sheet from Data_4.xlsx (output of Step 4)
1. Splits data by Type column (TRAIN / VALIDATE / TEST)
2. Trains 8 classifiers on TRAIN split, evaluates on all 3 phases
3. Saves ROC curve images in ./ROC_CURVES/
4. Performs LIME and SHAP analysis on TEST set for each classifier
5. Calculates Spearman correlation between SHAP and LIME rankings
6. Generates consensus features (top-5 in both SHAP and LIME)
7. Creates comprehensive Excel report with per-classifier and summary sheets
Input : Data_4.xlsx  (Data_Split sheet)
Output: Data_5.xlsx  +  ROC_CURVES/*.png
"""

# Imported modules
import os
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.stats import spearmanr

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sklearn.linear_model    import LogisticRegression
from sklearn.ensemble        import (RandomForestClassifier,
                                     GradientBoostingClassifier,
                                     ExtraTreesClassifier)
from sklearn.svm             import SVC
from sklearn.neighbors       import KNeighborsClassifier
from sklearn.tree            import DecisionTreeClassifier
from sklearn.naive_bayes     import GaussianNB
from sklearn.neural_network  import MLPClassifier
from sklearn.preprocessing   import StandardScaler
from sklearn.metrics         import (accuracy_score, precision_score,
                                     recall_score, f1_score,
                                     roc_auc_score, average_precision_score,
                                     roc_curve)

try:
    import lime
    import lime.lime_tabular
    LIME_AVAILABLE = True
except ImportError:
    print("⚠️  LIME not installed. Install with: pip install lime")
    LIME_AVAILABLE = False

try:
    import shap
    SHAP_AVAILABLE = True
except ImportError:
    print("⚠️  SHAP not installed. Install with: pip install shap")
    SHAP_AVAILABLE = False

warnings.filterwarnings('ignore')

# ========================== CONSTANTS ==========================

ROC_DIR         = 'ROC_CURVES'
PHASE_TRAIN     = 'TRAIN'
PHASE_VAL       = 'VALIDATE'
PHASE_TEST      = 'TEST'

# 8 classifiers (updated to match comparison)
CLASSIFIERS = {
    'Logistic Regression':   LogisticRegression(max_iter=2000, random_state=42),
    'Random Forest':         RandomForestClassifier(n_estimators=200, random_state=42, n_jobs=-1),
    'SVM (RBF Kernel)':      SVC(kernel='rbf', probability=True, random_state=42),
    'K-Nearest Neighbors':   KNeighborsClassifier(n_neighbors=5, n_jobs=-1),
    'Naive Bayes':           GaussianNB(),
    'Decision Tree':         DecisionTreeClassifier(random_state=42),
    'Gradient Boosting':     GradientBoostingClassifier(n_estimators=200, random_state=42),
    'Extra Trees':           ExtraTreesClassifier(n_estimators=200, random_state=42, n_jobs=-1),
}

PHASE_COLORS = {
    PHASE_TRAIN: '#2196F3',
    PHASE_VAL:   '#FF9800',
    PHASE_TEST:  '#4CAF50',
}

# ========================== DATA LOADER ==========================

class DataLoader:
    """Load Data_4.xlsx and split into train / validate / test."""

    def __init__(self, input_file='Data_4.xlsx'):
        self.input_file = input_file

    def load(self):
        print(f"\n{'='*80}")
        print(f"[STEP 0] LOADING DATA")
        print(f"{'='*80}")
        print(f"File : {self.input_file}")
        print(f"Sheet: Data_Split")

        df = pd.read_excel(self.input_file, sheet_name='Data_Split')

        print(f"\n✅ Loaded:")
        print(f"   Rows   : {len(df)}")
        print(f"   Columns: {len(df.columns)}")

        if 'Type' not in df.columns:
            raise ValueError("'Type' column not found in Data_Split sheet.")
        if 'Label' not in df.columns:
            raise ValueError("'Label' column not found in Data_Split sheet.")

        # Feature columns
        non_feature = {'Header', 'Sequence', 'Sequence_Length', 'Source',
                       'Label', 'Resistance_Probability', 'Type'}
        feature_cols = [c for c in df.columns if c not in non_feature]

        print(f"   Feature columns: {len(feature_cols)}")

        # Split by Type
        splits = {}
        for phase in [PHASE_TRAIN, PHASE_VAL, PHASE_TEST]:
            mask = df['Type'].str.upper() == phase
            sub  = df[mask].reset_index(drop=True)
            splits[phase] = (
                sub[feature_cols].astype(float).values,
                sub['Label'].astype(int).values,
            )
            print(f"   {phase:10s}: {len(sub)} rows")

        return splits, feature_cols


# ========================== METRICS CALCULATOR ==========================

class MetricsCalculator:
    """Compute 6 performance metrics for a phase."""

    @staticmethod
    def compute(y_true, y_pred, y_prob):
        return {
            'Accuracy':  accuracy_score(y_true, y_pred),
            'Precision': precision_score(y_true, y_pred, zero_division=0),
            'Recall':    recall_score(y_true, y_pred, zero_division=0),
            'F1_Score':  f1_score(y_true, y_pred, zero_division=0),
            'ROC_AUC':   roc_auc_score(y_true, y_prob) if len(np.unique(y_true)) > 1 else 0.0,
            'AUROC':     average_precision_score(y_true, y_prob) if len(np.unique(y_true)) > 1 else 0.0,
        }

# ========================== ROC CURVE PLOTTER ==========================

class ROCPlotter:
    """Generate ROC curves for each model."""

    def __init__(self, output_dir=ROC_DIR):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def plot(self, model_name, roc_data):
        fig, ax = plt.subplots(figsize=(8, 7))
        ax.set_facecolor('#F8F9FA')
        fig.patch.set_facecolor('white')

        for phase, data in roc_data.items():
            ax.plot(data['fpr'], data['tpr'],
                    color=PHASE_COLORS[phase],
                    linewidth=2.5,
                    label=f"{phase}  (AUC = {data['auc']:.4f})")

        ax.plot([0, 1], [0, 1], color='#BDBDBD', linewidth=1.2, linestyle='--',
                label='Random Classifier')

        ax.set_xlim([0.0, 1.0])
        ax.set_ylim([0.0, 1.05])
        ax.set_xlabel('False Positive Rate', fontsize=12, fontweight='bold')
        ax.set_ylabel('True Positive Rate', fontsize=12, fontweight='bold')
        ax.set_title(f'ROC Curve – {model_name}', fontsize=14, fontweight='bold', pad=20)
        ax.legend(loc='lower right', fontsize=11, framealpha=0.95)
        ax.grid(True, linestyle='--', alpha=0.5)

        path = os.path.join(self.output_dir, f'{model_name}_ROC.png')
        plt.tight_layout()
        plt.savefig(path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return path

# ========================== LIME EXPLAINER ==========================

class LIMEExplainer:
    """Generate LIME explanations and extract feature importance."""

    @staticmethod
    def explain(model_name, clf, X_test, y_test, feature_names, num_samples=1000):
        """
        Extract LIME feature importances for all test samples.
        Returns normalized importance scores.
        """
        if not LIME_AVAILABLE:
            print(f"  ⚠️  LIME unavailable for {model_name}")
            return None

        try:
            print(f"  🔄 Generating LIME explanations...")
            
            explainer = lime.lime_tabular.LimeTabularExplainer(
                training_data=X_test,
                feature_names=feature_names,
                class_names=['Not Resistant', 'Resistant'],
                mode='classification',
                random_state=42
            )

            # Get prediction function
            if hasattr(clf, 'predict_proba'):
                predict_fn = clf.predict_proba
            else:
                def predict_fn(X):
                    pred = clf.predict(X)
                    prob = np.zeros((len(pred), 2))
                    prob[np.arange(len(pred)), pred] = 1.0
                    return prob

            # Collect feature importances
            feature_importance = np.zeros(len(feature_names))
            num_explanations = min(100, len(X_test))

            for idx in range(num_explanations):
                try:
                    exp = explainer.explain_instance(X_test[idx], predict_fn, num_features=len(feature_names))
                    
                    for feat_name, weight in exp.as_list():
                        try:
                            # Extract feature name from LIME output
                            feat_clean = feat_name.split('<=')[0].split('>')[0].strip()
                            if feat_clean in feature_names:
                                f_idx = list(feature_names).index(feat_clean)
                                feature_importance[f_idx] += abs(weight)
                        except:
                            pass
                except:
                    pass

            # Normalize
            if feature_importance.max() > 0:
                feature_importance = feature_importance / feature_importance.max()
            
            return feature_importance

        except Exception as e:
            print(f"  ❌ LIME analysis failed for {model_name}: {str(e)}")
            return None

# ========================== SHAP EXPLAINER ==========================

class SHAPExplainer:
    """Generate SHAP explanations and extract feature importance."""

    @staticmethod
    def explain(model_name, clf, X_train_scaled, X_test_scaled, feature_names):
        """
        Extract SHAP feature importances.
        Returns normalized importance scores.
        """
        if not SHAP_AVAILABLE:
            print(f"  ⚠️  SHAP unavailable for {model_name}")
            return None

        try:
            print(f"  🔄 Generating SHAP explanations...")
            
            # Select background data
            background = X_train_scaled[:min(100, len(X_train_scaled))]
            
            # Create explainer
            if isinstance(clf, (RandomForestClassifier, ExtraTreesClassifier, 
                               GradientBoostingClassifier, DecisionTreeClassifier)):
                explainer = shap.TreeExplainer(clf)
            else:
                explainer = shap.KernelExplainer(
                    lambda x: clf.predict_proba(x)[:, 1],
                    shap.sample(background, 50)
                )
            
            # Calculate SHAP values
            shap_values = explainer.shap_values(X_test_scaled)
            
            # Handle different output formats
            if isinstance(shap_values, list):
                shap_vals = shap_values[1]  # Positive class
            else:
                shap_vals = shap_values
            
            # Mean absolute SHAP values
            mean_abs_shap = np.abs(shap_vals).mean(axis=0)
            
            # Normalize
            if mean_abs_shap.max() > 0:
                mean_abs_shap = mean_abs_shap / mean_abs_shap.max()
            
            return mean_abs_shap

        except Exception as e:
            print(f"  ❌ SHAP analysis failed for {model_name}: {str(e)}")
            return None

# ========================== FEATURE IMPORTANCE ANALYZER ==========================

class FeatureImportanceAnalyzer:
    """Compare LIME and SHAP rankings with Spearman correlation."""

    @staticmethod
    def analyze(model_name, lime_importance, shap_importance, feature_names, test_accuracy):
        """
        Compare LIME and SHAP importances.
        Returns a DataFrame with rankings and agreement metrics.
        """
        if lime_importance is None or shap_importance is None:
            print(f"  ⚠️  Cannot analyze {model_name}: missing LIME or SHAP")
            return None

        try:
            # Create ranking dataframe
            n_features = len(feature_names)
            shap_ranks = np.argsort(-shap_importance) + 1
            lime_ranks = np.argsort(-lime_importance) + 1
            
            # Spearman correlation
            spearman_rho, p_value = spearmanr(shap_ranks, lime_ranks)
            
            # Build results DataFrame
            results = pd.DataFrame({
                'Feature': feature_names,
                'SHAP_Importance': shap_importance,
                'SHAP_Rank': shap_ranks,
                'LIME_Importance': lime_importance,
                'LIME_Rank': lime_ranks,
                'Avg_Importance': (shap_importance + lime_importance) / 2,
            })
            
            # Sort by average importance (descending)
            results = results.sort_values('Avg_Importance', ascending=False).reset_index(drop=True)
            results['Consensus_Rank'] = np.arange(1, len(results) + 1)
            
            # Agreement classification
            results['SHAP_LIME_Agreement'] = results.apply(
                lambda row: _classify_agreement(row['SHAP_Rank'], row['LIME_Rank'], n_features),
                axis=1
            )
            
            return {
                'model_name': model_name,
                'test_accuracy': test_accuracy,
                'spearman_rho': spearman_rho,
                'results_df': results
            }

        except Exception as e:
            print(f"  ❌ Analysis failed for {model_name}: {str(e)}")
            return None


def _classify_agreement(shap_rank, lime_rank, n_features):
    """Classify agreement level between SHAP and LIME rankings."""
    diff = abs(shap_rank - lime_rank)
    threshold_high = n_features * 0.15
    threshold_medium = n_features * 0.30
    
    if diff <= threshold_high:
        return 'High'
    elif diff <= threshold_medium:
        return 'Medium'
    else:
        return 'Low'

# ========================== EXCEL WRITER ==========================

class ComprehensiveExcelWriter:
    """Write comprehensive LIME/SHAP analysis to Excel."""

    HEADER_FILL  = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11)
    YELLOW_FILL  = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    TITLE_FONT   = Font(bold=True, size=13)
    SECTION_FONT = Font(bold=True, size=11)
    BORDER       = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    @staticmethod
    def write(output_file, analysis_results, feature_cols):
        """
        Write comprehensive analysis to Excel with per-classifier and summary sheets.
        """
        print(f"\n{'='*80}")
        print(f"[EXCEL] Writing comprehensive analysis to {output_file}...")
        print(f"{'='*80}")

        wb = Workbook()
        wb.remove(wb.active)

        # ========== PER-CLASSIFIER SHEETS ==========
        classifier_data = {}
        for analysis in analysis_results:
            if analysis is None:
                continue
            
            model_name = analysis['model_name']
            test_accuracy = analysis['test_accuracy']
            spearman_rho = analysis['spearman_rho']
            results_df = analysis['results_df']
            
            classifier_data[model_name] = analysis
            
            # Create sheet
            ws = wb.create_sheet(model_name)
            
            # Title row
            title = f"Classifier: {model_name}"
            ws['A1'] = title
            ws['A1'].font = ComprehensiveExcelWriter.TITLE_FONT
            
            # Info row
            info = f"Test Accuracy: {test_accuracy:.2f}    |    SHAP↔LIME Spearman ρ = {spearman_rho:.4f}"
            ws['A2'] = info
            ws['A2'].font = Font(italic=True, size=10)
            
            # Empty row
            ws['A3'] = ""
            
            # Column headers
            headers = ['Feature', 'SHAP_Importance', 'SHAP_Rank', 'LIME_Importance', 
                      'LIME_Rank', 'Avg_Importance', 'Consensus_Rank', 'SHAP_LIME_Agreement']
            
            for col_idx, header in enumerate(headers, 1):
                c = ws.cell(4, col_idx, header)
                c.font = ComprehensiveExcelWriter.HEADER_FONT
                c.fill = ComprehensiveExcelWriter.HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = ComprehensiveExcelWriter.BORDER
            
            # Data rows
            for row_idx, (_, row) in enumerate(results_df.iterrows(), 5):
                ws.cell(row_idx, 1, row['Feature']).border = ComprehensiveExcelWriter.BORDER
                
                for col_idx, col_name in enumerate(['SHAP_Importance', 'SHAP_Rank', 'LIME_Importance',
                                                    'LIME_Rank', 'Avg_Importance', 'Consensus_Rank'], 2):
                    c = ws.cell(row_idx, col_idx, round(row[col_name], 6) if isinstance(row[col_name], float) else row[col_name])
                    c.number_format = '0.000000' if isinstance(row[col_name], float) else '0'
                    c.alignment = Alignment(horizontal='center')
                    c.border = ComprehensiveExcelWriter.BORDER
                
                # Agreement column
                c = ws.cell(row_idx, 8, row['SHAP_LIME_Agreement'])
                c.alignment = Alignment(horizontal='center')
                c.border = ComprehensiveExcelWriter.BORDER
                
                # Color code agreement
                if row['SHAP_LIME_Agreement'] == 'High':
                    c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif row['SHAP_LIME_Agreement'] == 'Medium':
                    c.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    c.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 20

        # ========== SUMMARY SHEET ==========
        if classifier_data:
            ws_summary = wb.create_sheet('Summary', 0)
            
            ws_summary['A1'] = 'SHAP + LIME Feature Importance — All Classifiers'
            ws_summary['A1'].font = ComprehensiveExcelWriter.TITLE_FONT
            
            ws_summary['A2'] = 'Sorted by Overall Score (mean of Mean_SHAP and Mean_LIME across all classifiers)'
            ws_summary['A2'].font = Font(italic=True, size=10)
            
            ws_summary['A3'] = ""
            
            # Build summary table
            summary_data = []
            feature_to_row = {}
            
            for idx, feature in enumerate(feature_cols, 1):
                row_data = {'Overall_Rank': None, 'Feature': feature}
                
                for model_name, analysis in classifier_data.items():
                    results_df = analysis['results_df']
                    feat_data = results_df[results_df['Feature'] == feature]
                    
                    if not feat_data.empty:
                        row_data[f'{model_name}_SHAP'] = feat_data['SHAP_Importance'].values[0]
                        row_data[f'{model_name}_LIME'] = feat_data['LIME_Importance'].values[0]
                    else:
                        row_data[f'{model_name}_SHAP'] = 0
                        row_data[f'{model_name}_LIME'] = 0
                
                summary_data.append(row_data)
                feature_to_row[feature] = idx
            
            # Calculate mean SHAP and LIME
            summary_df = pd.DataFrame(summary_data)
            
            shap_cols = [c for c in summary_df.columns if '_SHAP' in c]
            lime_cols = [c for c in summary_df.columns if '_LIME' in c]
            
            summary_df['Mean_SHAP'] = summary_df[shap_cols].mean(axis=1)
            summary_df['Mean_LIME'] = summary_df[lime_cols].mean(axis=1)
            summary_df['Overall_Score'] = (summary_df['Mean_SHAP'] + summary_df['Mean_LIME']) / 2
            
            # Sort by overall score
            summary_df = summary_df.sort_values('Overall_Score', ascending=False).reset_index(drop=True)
            summary_df['Overall_Rank'] = np.arange(1, len(summary_df) + 1)
            
            # Reorder columns
            cols = ['Overall_Rank', 'Feature'] + shap_cols + lime_cols + ['Mean_SHAP', 'Mean_LIME', 'Overall_Score']
            summary_df = summary_df[cols]
            
            # Write to sheet
            headers = summary_df.columns.tolist()
            for col_idx, header in enumerate(headers, 1):
                c = ws_summary.cell(4, col_idx, header)
                c.font = ComprehensiveExcelWriter.HEADER_FONT
                c.fill = ComprehensiveExcelWriter.HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = ComprehensiveExcelWriter.BORDER
            
            # Data rows
            for row_idx, (_, row) in enumerate(summary_df.iterrows(), 5):
                for col_idx, col_name in enumerate(headers, 1):
                    c = ws_summary.cell(row_idx, col_idx, row[col_name])
                    
                    if col_name in ['Overall_Rank', 'Feature']:
                        c.alignment = Alignment(horizontal='left')
                    else:
                        c.number_format = '0.000000'
                        c.alignment = Alignment(horizontal='center')
                    
                    c.border = ComprehensiveExcelWriter.BORDER
            
            # Column widths
            ws_summary.column_dimensions['A'].width = 14
            ws_summary.column_dimensions['B'].width = 15
            for col_idx in range(3, len(headers) + 1):
                ws_summary.column_dimensions[get_column_letter(col_idx)].width = 16

        # ========== RANK CORRELATION SHEET ==========
        ws_corr = wb.create_sheet('Rank_Correlation')
        
        ws_corr['A1'] = 'Spearman ρ between SHAP and LIME per Classifier'
        ws_corr['A1'].font = ComprehensiveExcelWriter.TITLE_FONT
        
        ws_corr['A2'] = ""
        
        headers_corr = ['Classifier', 'Test_Accuracy', 'SHAP_LIME_Spearman', 'Interpretation']
        for col_idx, header in enumerate(headers_corr, 1):
            c = ws_corr.cell(3, col_idx, header)
            c.font = ComprehensiveExcelWriter.HEADER_FONT
            c.fill = ComprehensiveExcelWriter.HEADER_FILL
            c.border = ComprehensiveExcelWriter.BORDER
        
        row_idx = 4
        for model_name, analysis in classifier_data.items():
            spearman_rho = analysis['spearman_rho']
            test_acc = analysis['test_accuracy']
            
            # Interpretation
            if spearman_rho > 0.8:
                interp = "Strong agreement"
            elif spearman_rho > 0.5:
                interp = "Moderate agreement"
            else:
                interp = "Weak agreement"
            
            ws_corr.cell(row_idx, 1, model_name).border = ComprehensiveExcelWriter.BORDER
            ws_corr.cell(row_idx, 2, round(test_acc, 4)).number_format = '0.0000'
            ws_corr.cell(row_idx, 2).border = ComprehensiveExcelWriter.BORDER
            ws_corr.cell(row_idx, 3, round(spearman_rho, 4)).number_format = '0.0000'
            ws_corr.cell(row_idx, 3).border = ComprehensiveExcelWriter.BORDER
            ws_corr.cell(row_idx, 4, interp).border = ComprehensiveExcelWriter.BORDER
            
            row_idx += 1
        
        ws_corr.column_dimensions['A'].width = 25
        ws_corr.column_dimensions['B'].width = 16
        ws_corr.column_dimensions['C'].width = 22
        ws_corr.column_dimensions['D'].width = 20

        # ========== CONSENSUS SHEET ==========
        ws_consensus = wb.create_sheet('Consensus')
        
        ws_consensus['A1'] = 'Consensus — features in top-5 SHAP AND top-5 LIME across classifiers'
        ws_consensus['A1'].font = ComprehensiveExcelWriter.TITLE_FONT
        
        ws_consensus['A2'] = ""
        
        # Find consensus features
        consensus_features = {}
        for model_name, analysis in classifier_data.items():
            results_df = analysis['results_df']
            top5_shap = set(results_df[results_df['SHAP_Rank'] <= 5]['Feature'].tolist())
            top5_lime = set(results_df[results_df['LIME_Rank'] <= 5]['Feature'].tolist())
            consensus = top5_shap & top5_lime
            
            for feature in consensus:
                if feature not in consensus_features:
                    consensus_features[feature] = 0
                consensus_features[feature] += 1
        
        # Sort by agreement count
        consensus_list = sorted(consensus_features.items(), key=lambda x: -x[1])
        
        headers_consensus = ['Feature', 'Classifiers_Agreed', 'Overall_Rank']
        for col_idx, header in enumerate(headers_consensus, 1):
            c = ws_consensus.cell(3, col_idx, header)
            c.font = ComprehensiveExcelWriter.HEADER_FONT
            c.fill = ComprehensiveExcelWriter.HEADER_FILL
            c.border = ComprehensiveExcelWriter.BORDER
        
        for rank, (feature, count) in enumerate(consensus_list, 1):
            ws_consensus.cell(3 + rank, 1, feature).border = ComprehensiveExcelWriter.BORDER
            ws_consensus.cell(3 + rank, 2, count).border = ComprehensiveExcelWriter.BORDER
            ws_consensus.cell(3 + rank, 3, rank).border = ComprehensiveExcelWriter.BORDER
        
        ws_consensus.column_dimensions['A'].width = 15
        ws_consensus.column_dimensions['B'].width = 18
        ws_consensus.column_dimensions['C'].width = 15

        wb.save(output_file)
        print(f"✅ Saved: {output_file}")

# ========================== MODEL TRAINER ==========================

class ComprehensiveModelTrainer:
    """Train models and perform comprehensive LIME/SHAP analysis."""

    def __init__(self):
        self.roc_plotter = ROCPlotter()
        self.scaler = StandardScaler()
        self.lime_explainer = LIMEExplainer()
        self.shap_explainer = SHAPExplainer()
        self.analyzer = FeatureImportanceAnalyzer()
        self.analysis_results = []

    def run(self, splits, feature_cols):
        """Train models and analyze with LIME/SHAP."""
        X_train, y_train = splits[PHASE_TRAIN]

        # Scale data
        X_train_sc = self.scaler.fit_transform(X_train)
        scaled = {
            PHASE_TRAIN: (X_train_sc, y_train),
            PHASE_VAL:   (self.scaler.transform(splits[PHASE_VAL][0]), splits[PHASE_VAL][1]),
            PHASE_TEST:  (self.scaler.transform(splits[PHASE_TEST][0]), splits[PHASE_TEST][1]),
        }

        X_test_sc, y_test = scaled[PHASE_TEST]
        all_rows = []
        phase_order = [PHASE_TRAIN, PHASE_VAL, PHASE_TEST]

        for model_name, clf in CLASSIFIERS.items():
            print(f"\n{'='*80}")
            print(f"[MODEL] {model_name}")
            print(f"{'='*80}")

            # Train
            clf.fit(X_train_sc, y_train)
            print(f"  Fitted on {len(y_train)} training samples.")

            roc_data = {}
            test_accuracy = None

            for phase in phase_order:
                X_ph, y_ph = scaled[phase]
                y_pred = clf.predict(X_ph)

                # Probability scores
                if hasattr(clf, 'predict_proba'):
                    y_prob = clf.predict_proba(X_ph)[:, 1]
                elif hasattr(clf, 'decision_function'):
                    raw    = clf.decision_function(X_ph)
                    y_prob = (raw - raw.min()) / (raw.max() - raw.min() + 1e-9)
                else:
                    y_prob = y_pred.astype(float)

                metrics = MetricsCalculator.compute(y_ph, y_pred, y_prob)

                row = {'Model': model_name, 'Phase': phase}
                row.update(metrics)
                all_rows.append(row)

                print(f"  {phase:10s} | Acc={metrics['Accuracy']:.4f}  F1={metrics['F1_Score']:.4f}  "
                      f"ROC_AUC={metrics['ROC_AUC']:.4f}")

                if phase == PHASE_TEST:
                    test_accuracy = metrics['Accuracy']

                # ROC data
                if len(np.unique(y_ph)) > 1:
                    fpr, tpr, _ = roc_curve(y_ph, y_prob)
                else:
                    fpr = tpr = np.array([0.0, 1.0])

                roc_data[phase] = {
                    'fpr': fpr,
                    'tpr': tpr,
                    'auc': metrics['ROC_AUC'],
                }

            # ROC plot
            img_path = self.roc_plotter.plot(model_name, roc_data)
            print(f"  ROC curve → {img_path}")

            # LIME analysis
            lime_importance = self.lime_explainer.explain(
                model_name, clf, X_test_sc, y_test, feature_cols
            )

            # SHAP analysis
            shap_importance = self.shap_explainer.explain(
                model_name, clf, X_train_sc, X_test_sc, feature_cols
            )

            # Compare LIME and SHAP
            analysis = self.analyzer.analyze(
                model_name, lime_importance, shap_importance, feature_cols, test_accuracy
            )
            
            if analysis:
                self.analysis_results.append(analysis)
                print(f"  ✅ LIME/SHAP analysis complete")
                print(f"     Spearman ρ = {analysis['spearman_rho']:.4f}")

        results_df = pd.DataFrame(all_rows)
        
        # Find best model by test F1
        test_rows = results_df[results_df['Phase'] == PHASE_TEST].copy()
        best_idx = test_rows['F1_Score'].idxmax()
        best_model = test_rows.loc[best_idx, 'Model']

        print(f"\n{'='*80}")
        print(f"✅ Best Model: {best_model} (Test F1 = {test_rows.loc[best_idx, 'F1_Score']:.4f})")
        print(f"{'='*80}\n")

        return results_df, best_model

# ========================== MAIN PIPELINE ==========================

class ComprehensivePipeline:
    """Orchestrate comprehensive LIME/SHAP analysis."""

    def __init__(self, input_file='Data_4.xlsx', output_file='Data_5_Comprehensive.xlsx'):
        self.input_file = input_file
        self.output_file = output_file

    def run(self):
        print("\n" + "="*80)
        print("STEP 5: COMPREHENSIVE LIME & SHAP ANALYSIS")
        print("Per-Classifier Analysis | Spearman Correlation | Consensus Features")
        print("="*80)

        # Load data
        loader = DataLoader(self.input_file)
        splits, feature_cols = loader.load()

        # Train and analyze
        trainer = ComprehensiveModelTrainer()
        results_df, best_model = trainer.run(splits, feature_cols)

        # Write comprehensive Excel
        ComprehensiveExcelWriter.write(self.output_file, trainer.analysis_results, feature_cols)

        print(f"\n{'='*80}")
        print("✅ COMPLETE: Comprehensive LIME/SHAP analysis finished!")
        print(f"   Output → {self.output_file}")
        print(f"   Sheets: Classifier-specific + Summary + Rank_Correlation + Consensus")
        print(f"   ROC plots → ./{ROC_DIR}/")
        print(f"{'='*80}\n")


# ====================== ENTRY POINT ======================

if __name__ == "__main__":
    pipeline = ComprehensivePipeline(
        input_file='Data_4.xlsx',
        output_file='Data_5_Comprehensive.xlsx',
    )
    pipeline.run()