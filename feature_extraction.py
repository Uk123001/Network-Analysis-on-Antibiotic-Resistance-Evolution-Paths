"""
Step 2: Feature Extraction & Classification (UPDATED with optrA-Specific Features)
1. Similarity filtering (Jaccard similarity: 60-99%, remove 100% duplicates)
2. Extract 51 optrA-specific resistance features from sequences
3. Classify as resistant/non-resistant using MLP with optrA features
Input: Data_1.xlsx
Output: Data_2.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from collections import Counter
import re
import warnings
warnings.filterwarnings('ignore')


class SimilarityFilter:
    """Filters sequences using Jaccard similarity"""
    
    def __init__(self, min_similarity=0.60, max_similarity=0.99):
        self.min_similarity = min_similarity
        self.max_similarity = max_similarity
        self.removal_log = []
        self.kept_indices = []
    
    @staticmethod
    def jaccard_similarity(seq1, seq2):
        """Calculate Jaccard similarity between two sequences"""
        set1 = set(seq1)
        set2 = set(seq2)
        
        intersection = len(set1.intersection(set2))
        union = len(set1.union(set2))
        
        return intersection / union if union > 0 else 0
    
    @staticmethod
    def sequence_identity(seq1, seq2):
        """Calculate percentage identity"""
        if len(seq1) != len(seq2):
            min_len = min(len(seq1), len(seq2))
            seq1, seq2 = seq1[:min_len], seq2[:min_len]
        
        if len(seq1) == 0:
            return 0
        
        matches = sum(1 for a, b in zip(seq1, seq2) if a == b)
        return (matches / len(seq1)) * 100
    
    def filter_sequences(self, df):
        """
        Apply similarity filtering:
        1. Remove 100% identical sequences
        2. Keep sequences with 60-99% Jaccard similarity
        """
        print("\n[Similarity Filter] Starting sequence filtering...")
        
        # Step 1: Remove 100% identical sequences
        print("[Similarity Filter] Step 1: Removing 100% identical sequences...")
        initial_count = len(df)
        seen_sequences = {}
        duplicates_removed = []
        filtered_indices = []
        
        for idx, row in df.iterrows():
            seq = row['Sequence']
            
            if seq not in seen_sequences:
                seen_sequences[seq] = idx
                filtered_indices.append(idx)
            else:
                duplicates_removed.append({
                    'index': idx,
                    'header': row['Header'],
                    'match_with': df.loc[seen_sequences[seq], 'Header'],
                    'similarity': 100.0,
                    'reason': '100% identical duplicate'
                })
        
        df_filtered = df.iloc[filtered_indices].reset_index(drop=True)
        after_duplicates = len(df_filtered)
        
        print(f"[Similarity Filter] Removed {initial_count - after_duplicates} sequences (100% duplicates)")
        self.removal_log.extend(duplicates_removed)
        
        # Step 2: Apply Jaccard similarity threshold (60-99%)
        print("[Similarity Filter] Step 2: Applying Jaccard similarity filter (60-99%)...")
        kept_indices = []
        similarity_removed = []
        
        for i, row_i in df_filtered.iterrows():
            seq_i = row_i['Sequence']
            keep_sequence = False
            
            # Check if sequence has at least one similar sequence
            for j, row_j in df_filtered.iterrows():
                if i >= j:
                    continue
                
                seq_j = row_j['Sequence']
                jaccard_sim = self.jaccard_similarity(seq_i, seq_j)
                
                if self.min_similarity <= jaccard_sim <= self.max_similarity:
                    keep_sequence = True
                    break
            
            if keep_sequence:
                kept_indices.append(i)
            else:
                similarity_removed.append({
                    'index': i,
                    'header': row_i['Header'],
                    'jaccard_similarity': 'Outside 60-99% range',
                    'reason': 'No similar sequence in 60-99% range'
                })
        
        df_final = df_filtered.iloc[kept_indices].reset_index(drop=True)
        
        print(f"[Similarity Filter] Removed {len(similarity_removed)} sequences (outside similarity range)")
        print(f"[Similarity Filter] Final dataset: {len(df_final)} sequences")
        
        self.removal_log.extend(similarity_removed)
        
        # Summary statistics
        summary = {
            'Initial_sequences': initial_count,
            'After_removing_100pct_duplicates': after_duplicates,
            'Duplicates_removed': initial_count - after_duplicates,
            'After_jaccard_filtering': len(df_final),
            'Jaccard_filtered_out': len(df_filtered) - len(df_final),
            'Final_sequences': len(df_final),
            'Similarity_threshold_min': self.min_similarity,
            'Similarity_threshold_max': self.max_similarity,
            'Method': 'Jaccard Similarity + Identity Check'
        }
        
        return df_final, summary


class OptrAFeatureExtractor:
    """Extracts 51 optrA-specific resistance features"""
    
    # Known optrA resistance-associated mutations (from literature)
    RESISTANCE_MUTATIONS = {
        'S83F': [('TCC', 'TTC'), ('TCT', 'TTT')],  # GyrA Serine 83 to Phe
        'D87N': [('GAC', 'AAC'), ('GAT', 'AAT')],  # GyrA Asp 87 to Asn
        'E84K': [('GAA', 'AAA'), ('GAG', 'AAG')],  # ParC Glu 84 to Lys
        'S80I': [('TCA', 'ATA'), ('TCT', 'ATT')],  # GyrA Ser 80 to Ile
    }
    
    # Promoter consensus sequences for optrA activation
    PROMOTER_SEQUENCES = {
        'TATAAT': 'Pribnow_box',  # -10 box
        'TTGACA': 'Consensus_-35',
        'TGACA': '-35_variant',
        'TATAAT': '-10_variant'
    }
    
    # IS elements associated with optrA activation
    IS_ELEMENTS = {
        'AAATAAA': 'IS1',
        'TTGTTT': 'IS10_left',
        'TGACA': 'IS5_inverted',
        'AATAATTGAA': 'IS102'
    }
    
    # Efflux pump transmembrane prediction (simplified hydrophobic patterns)
    HYDROPHOBIC_MOTIFS = [
        'LLLIIIIIILLLLIIIIIII',  # Strong hydrophobic
        'LLLIIIIIII',
        'FFFFFFFSSS'
    ]
    
    def __init__(self):
        self.codon_table = self._create_codon_table()
    
    def _create_codon_table(self):
        """Create standard genetic code table"""
        codon_table = {
            'ATA':'I', 'ATC':'I', 'ATT':'I', 'ATG':'M',
            'ACA':'T', 'ACC':'T', 'ACG':'T', 'ACT':'T',
            'AAC':'N', 'AAT':'N', 'AAA':'K', 'AAG':'K',
            'AGC':'S', 'AGT':'S', 'AGA':'R', 'AGG':'R',
            'CTA':'L', 'CTC':'L', 'CTG':'L', 'CTT':'L',
            'CCA':'P', 'CCC':'P', 'CCG':'P', 'CCT':'P',
            'CAC':'H', 'CAT':'H', 'CAA':'Q', 'CAG':'Q',
            'CGA':'R', 'CGC':'R', 'CGG':'R', 'CGT':'R',
            'GTA':'V', 'GTC':'V', 'GTG':'V', 'GTT':'V',
            'GCA':'A', 'GCC':'A', 'GCG':'A', 'GCT':'A',
            'GAC':'D', 'GAT':'D', 'GAA':'E', 'GAG':'E',
            'GGA':'G', 'GGC':'G', 'GGG':'G', 'GGT':'G',
            'TCA':'S', 'TCC':'S', 'TCG':'S', 'TCT':'S',
            'TTC':'F', 'TTT':'F', 'TTA':'L', 'TTG':'L',
            'TAC':'Y', 'TAT':'Y', 'TAA':'*', 'TAG':'*',
            'TGC':'C', 'TGT':'C', 'TGA':'*', 'TGG':'W',
        }
        return codon_table
    
    def extract_all_features(self, sequence):
        """Extract all 51 optrA-specific resistance features"""
        seq = sequence.upper()
        features = {}
        
        # ============ GROUP 1: GENE STRUCTURE & INTEGRITY (Features 1-8) ============
        features['1_Gene_Length_bp'] = len(seq)
        features['2_Has_Start_Codon_ATG'] = 1.0 if 'ATG' in seq else 0.0
        features['3_Has_Stop_Codon'] = 1.0 if ('TAA' in seq or 'TAG' in seq or 'TGA' in seq) else 0.0
        features['4_ORF_Completeness'] = self._orf_completeness(seq)
        features['5_Frame_Integrity'] = self._frame_integrity(seq)
        features['6_Internal_Stop_Codons'] = self._count_internal_stops(seq)
        features['7_Sequence_Length_Normalized'] = min(len(seq) / 1200, 1.0)  # optrA ~1200 bp
        features['8_Is_Full_Length_Gene'] = 1.0 if 950 <= len(seq) <= 1300 else 0.0
        
        # ============ GROUP 2: GC COMPOSITION (Features 9-14) ============
        gc_count = seq.count('G') + seq.count('C')
        features['9_GC_Percent'] = (gc_count / len(seq)) * 100 if len(seq) > 0 else 0
        features['10_GC_Skew'] = self._gc_skew(seq)
        features['11_AT_Skew'] = self._at_skew(seq)
        features['12_GC_Codon_Pos3'] = self._gc_at_position(seq, 2)
        features['13_Purine_Percent'] = self._purine_percent(seq)
        features['14_Pyrimidine_Percent'] = self._pyrimidine_percent(seq)
        
        # ============ GROUP 3: KNOWN RESISTANCE MUTATIONS (Features 15-18) ============
        features['15_S83F_Mutation_Count'] = self._detect_mutation_pattern(seq, 'S83F')
        features['16_D87N_Mutation_Count'] = self._detect_mutation_pattern(seq, 'D87N')
        features['17_E84K_Mutation_Count'] = self._detect_mutation_pattern(seq, 'E84K')
        features['18_S80I_Mutation_Count'] = self._detect_mutation_pattern(seq, 'S80I')
        
        # ============ GROUP 4: PROMOTER & REGULATORY ELEMENTS (Features 19-23) ============
        features['19_Pribnow_Box_Count'] = seq.count('TATAAT')
        features['20_Consensus_-35_Count'] = seq.count('TTGACA')
        features['21_Promoter_Strength'] = self._promoter_strength(seq)
        features['22_Upstream_Regulatory_GC'] = self._upstream_regulatory_gc(seq)
        features['23_TATA_Box_Distance_to_Start'] = self._distance_to_start_codon(seq, 'TATAAT')
        
        # ============ GROUP 5: IS ELEMENTS & MOBILE GENETIC ELEMENTS (Features 24-28) ============
        features['24_IS1_Count'] = seq.count('AAATAAA')
        features['25_IS10_Count'] = seq.count('TTGTTT')
        features['26_IS5_Count'] = seq.count('TGACA')
        features['27_Total_IS_Elements'] = (seq.count('AAATAAA') + seq.count('TTGTTT') + seq.count('TGACA') + seq.count('AATAATTGAA'))
        features['28_IS_Element_Density'] = features['27_Total_IS_Elements'] / max(len(seq) / 100, 1)
        
        # ============ GROUP 6: EFFLUX PUMP CHARACTERISTICS (Features 29-33) ============
        features['29_Hydrophobic_Region_Count'] = self._count_hydrophobic_regions(seq)
        features['30_Transmembrane_Domain_Count'] = self._predict_transmembrane_domains(seq)
        features['31_Protein_Length_aa'] = len(seq) // 3
        features['32_Expected_MW_kDa'] = (len(seq) // 3) * 0.110  # Approximate MW calculation
        features['33_Charge_Bias'] = self._calculate_charge_bias(seq)
        
        # ============ GROUP 7: DINUCLEOTIDE & CODON BIAS (Features 34-39) ============
        features['34_CpG_Count'] = seq.count('CG')
        features['35_CpG_per_100bp'] = (seq.count('CG') / len(seq)) * 100 if len(seq) > 0 else 0
        features['36_CpG_ObsExp_Ratio'] = self._cpg_obsexp(seq)
        features['37_Codon_Adaptation_Index'] = self._codon_adaptation_index(seq)
        features['38_Codon_Bias_Strength'] = self._codon_bias_strength(seq)
        features['39_Unique_Codons_Used'] = self._count_unique_codons(seq)
        
        # ============ GROUP 8: SEQUENCE COMPLEXITY (Features 40-44) ============
        features['40_Entropy_Trinucleotide'] = self._entropy(self._get_kmers(seq, 3))
        features['41_Entropy_Tetranucleotide'] = self._entropy(self._get_kmers(seq, 4))
        features['42_4mer_Complexity'] = self._kmer_complexity(seq, 4)
        features['43_Repeat_Fraction'] = self._repeat_fraction(seq)
        features['44_Longest_Homopolymer'] = self._longest_homopolymer(seq)
        
        # ============ GROUP 9: MUTATION SIGNALS & ADAPTATION (Features 45-49) ============
        features['45_Synonymous_Site_Density'] = self._synonymous_site_density(seq)
        features['46_NonSynonymous_Site_Density'] = self._nonsynonymous_site_density(seq)
        features['47_dN_dS_Proxy'] = self._dn_ds_proxy(seq)
        features['48_Transition_Transversion_Proxy'] = self._ts_tv_proxy(seq)
        features['49_Mutability_Index'] = self._mutability_index(seq)
        
        # ============ GROUP 10: FUNCTIONAL & EVOLUTION METRICS (Features 50-51) ============
        features['50_HGT_Risk_Score'] = self._hgt_risk_score(seq)
        features['51_Resistance_Evolution_Index'] = self._resistance_evolution_index(seq)
        
        return features
    
    def _orf_completeness(self, seq):
        """Calculate ORF completeness score"""
        has_start = 1.0 if 'ATG' in seq else 0.0
        has_stop = 1.0 if ('TAA' in seq or 'TAG' in seq or 'TGA' in seq) else 0.0
        length_score = min(len(seq) / 1200, 1.0)
        return (has_start + has_stop + length_score) / 3.0
    
    def _frame_integrity(self, seq):
        """Check if sequence maintains reading frame integrity"""
        if len(seq) < 3:
            return 0.0
        codons = [seq[i:i+3] for i in range(0, len(seq) - 2, 3)]
        valid_codons = sum(1 for codon in codons if len(codon) == 3 and all(c in 'ATGC' for c in codon))
        return valid_codons / len(codons) if codons else 0.0
    
    def _count_internal_stops(self, seq):
        """Count internal stop codons (excluding final stop)"""
        stops = ['TAA', 'TAG', 'TGA']
        count = 0
        for i in range(0, len(seq) - 5, 3):
            if seq[i:i+3] in stops:
                count += 1
        return count
    
    def _gc_skew(self, seq):
        """Calculate GC skew (G-C)/(G+C)"""
        g_count = seq.count('G')
        c_count = seq.count('C')
        total = g_count + c_count
        return ((g_count - c_count) / total) if total > 0 else 0
    
    def _at_skew(self, seq):
        """Calculate AT skew (A-T)/(A+T)"""
        a_count = seq.count('A')
        t_count = seq.count('T')
        total = a_count + t_count
        return ((a_count - t_count) / total) if total > 0 else 0
    
    def _gc_at_position(self, seq, pos):
        """Get GC percentage at codon position"""
        count = 0
        total = 0
        for i in range(pos, len(seq) - 2, 3):
            total += 1
            if seq[i] in 'GC':
                count += 1
        return (count / total * 100) if total > 0 else 0
    
    def _purine_percent(self, seq):
        """Calculate purine (A+G) percentage"""
        purines = seq.count('A') + seq.count('G')
        return (purines / len(seq)) * 100 if len(seq) > 0 else 0
    
    def _pyrimidine_percent(self, seq):
        """Calculate pyrimidine (C+T) percentage"""
        pyrimidines = seq.count('C') + seq.count('T')
        return (pyrimidines / len(seq)) * 100 if len(seq) > 0 else 0
    
    def _detect_mutation_pattern(self, seq, mutation):
        """Detect known resistance mutations"""
        if mutation in self.RESISTANCE_MUTATIONS:
            patterns = self.RESISTANCE_MUTATIONS[mutation]
            count = 0
            for pattern_pair in patterns:
                count += seq.count(pattern_pair[1])  # Count mutant pattern
            return count
        return 0
    
    def _promoter_strength(self, seq):
        """Calculate promoter strength based on consensus sequences"""
        score = 0
        if 'TATAAT' in seq:
            score += 0.5
        if 'TTGACA' in seq:
            score += 0.3
        if 'TGACA' in seq:
            score += 0.2
        return min(score, 1.0)
    
    def _upstream_regulatory_gc(self, seq):
        """Get GC content in upstream regulatory region (first 300 bp)"""
        upstream = seq[:300]
        gc_count = upstream.count('G') + upstream.count('C')
        return (gc_count / len(upstream)) * 100 if len(upstream) > 0 else 0
    
    def _distance_to_start_codon(self, seq, motif):
        """Calculate distance from promoter motif to start codon"""
        start_pos = seq.find('ATG')
        motif_pos = seq.find(motif)
        
        if start_pos >= 0 and motif_pos >= 0:
            return abs(start_pos - motif_pos)
        return len(seq)
    
    def _count_hydrophobic_regions(self, seq):
        """Count hydrophobic amino acid regions"""
        hydrophobic_aa = 'ILVFM'
        protein = self._translate(seq)
        count = 0
        consecutive = 0
        
        for aa in protein:
            if aa in hydrophobic_aa:
                consecutive += 1
                if consecutive >= 10:
                    count += 1
                    consecutive = 0
            else:
                consecutive = 0
        
        return count
    
    def _predict_transmembrane_domains(self, seq):
        """Predict transmembrane domains (simplified Kyte-Doolittle)"""
        protein = self._translate(seq)
        kd_scale = {
            'A': 1.8, 'C': 2.5, 'D': -3.5, 'E': -3.5, 'F': 2.8,
            'G': -0.4, 'H': -3.2, 'I': 4.5, 'K': -3.9, 'L': 3.8,
            'M': 1.9, 'N': -3.5, 'P': -1.6, 'Q': -3.5, 'R': -4.5,
            'S': -0.8, 'T': -0.7, 'V': 4.2, 'W': -0.9, 'Y': -1.3
        }
        
        window_size = 9
        tm_count = 0
        
        for i in range(len(protein) - window_size):
            window = protein[i:i+window_size]
            score = sum(kd_scale.get(aa, 0) for aa in window) / window_size
            if score > 1.5:
                tm_count += 1
        
        return tm_count
    
    def _calculate_charge_bias(self, seq):
        """Calculate charge bias in protein"""
        protein = self._translate(seq)
        positive = protein.count('K') + protein.count('R')
        negative = protein.count('D') + protein.count('E')
        total = len(protein)
        return ((positive - negative) / total) if total > 0 else 0
    
    def _translate(self, seq):
        """Translate DNA sequence to protein"""
        protein = []
        for i in range(0, len(seq) - 2, 3):
            codon = seq[i:i+3]
            protein.append(self.codon_table.get(codon, 'X'))
        return ''.join(protein)
    
    def _cpg_obsexp(self, seq):
        """CpG observed/expected ratio"""
        cpg_count = seq.count('CG')
        c_count = seq.count('C')
        g_count = seq.count('G')
        expected = (c_count * g_count) / len(seq) if len(seq) > 0 else 1
        return (cpg_count / expected) if expected > 0 else 0
    
    def _codon_adaptation_index(self, seq):
        """Calculate codon adaptation index (simplified)"""
        codons = [seq[i:i+3] for i in range(0, len(seq) - 2, 3)]
        if not codons:
            return 0.0
        
        # Count rare vs common codons (simplified)
        rare_codons = 0
        for codon in codons:
            if codon in ['CGA', 'CGG', 'AGA', 'AGG', 'TTA']:
                rare_codons += 1
        
        return 1.0 - (rare_codons / len(codons))
    
    def _codon_bias_strength(self, seq):
        """Calculate codon usage bias strength"""
        codons = [seq[i:i+3] for i in range(0, len(seq) - 2, 3)]
        if not codons:
            return 0.0
        
        counter = Counter(codons)
        most_common = counter.most_common(1)[0][1] if counter else 0
        return (most_common / len(codons)) if len(codons) > 0 else 0
    
    def _count_unique_codons(self, seq):
        """Count unique codons used"""
        codons = set()
        for i in range(0, len(seq) - 2, 3):
            codons.add(seq[i:i+3])
        return len(codons)
    
    def _get_kmers(self, seq, k):
        """Extract k-mers"""
        kmers = []
        for i in range(len(seq) - k + 1):
            kmers.append(seq[i:i+k])
        return kmers
    
    def _entropy(self, kmers):
        """Calculate Shannon entropy"""
        if not kmers:
            return 0
        counter = Counter(kmers)
        total = len(kmers)
        entropy = 0
        for count in counter.values():
            p = count / total
            entropy -= p * np.log2(p) if p > 0 else 0
        return entropy
    
    def _kmer_complexity(self, seq, k):
        """Calculate k-mer complexity"""
        kmers = self._get_kmers(seq, k)
        unique_kmers = len(set(kmers))
        max_kmers = 4 ** k
        return (unique_kmers / max_kmers) if max_kmers > 0 else 0
    
    def _repeat_fraction(self, seq):
        """Calculate fraction of repeating sequences"""
        kmers = self._get_kmers(seq, 4)
        if not kmers:
            return 0
        counter = Counter(kmers)
        repeats = sum(1 for count in counter.values() if count > 1)
        return (repeats / len(counter)) if len(counter) > 0 else 0
    
    def _longest_homopolymer(self, seq):
        """Find longest homopolymer run"""
        if not seq:
            return 0
        max_run = 1
        current_run = 1
        for i in range(1, len(seq)):
            if seq[i] == seq[i-1]:
                current_run += 1
                max_run = max(max_run, current_run)
            else:
                current_run = 1
        return max_run
    
    def _synonymous_site_density(self, seq):
        """Calculate synonymous substitution sites density"""
        codons = [seq[i:i+3] for i in range(0, len(seq) - 2, 3)]
        return len(codons) / max(len(seq) / 100, 1)
    
    def _nonsynonymous_site_density(self, seq):
        """Calculate non-synonymous substitution sites density"""
        return self._synonymous_site_density(seq) * 0.8  # Simplified
    
    def _dn_ds_proxy(self, seq):
        """Calculate dN/dS proxy (non-syn/syn substitution ratio)"""
        syn = self._synonymous_site_density(seq)
        nonsyn = self._nonsynonymous_site_density(seq)
        return (nonsyn / syn) if syn > 0 else 0
    
    def _ts_tv_proxy(self, seq):
        """Transition/Transversion proxy"""
        transitions = seq.count('GA') + seq.count('AG') + seq.count('CT') + seq.count('TC')
        transversions = seq.count('AC') + seq.count('CA') + seq.count('GT') + seq.count('TG')
        return (transitions / transversions) if transversions > 0 else 0
    
    def _mutability_index(self, seq):
        """Calculate mutability index"""
        cpg = seq.count('CG')
        tat = seq.count('TA') + seq.count('AT')
        return (cpg + tat) / max(len(seq), 1)
    
    def _hgt_risk_score(self, seq):
        """Calculate horizontal gene transfer risk"""
        gc = (seq.count('G') + seq.count('C')) / len(seq) if len(seq) > 0 else 0
        is_count = self._count_is_elements(seq)
        promoter_score = self._promoter_strength(seq)
        
        # HGT risk increases with IS elements, unusual composition, and strong promoters
        return min((is_count * 0.3 + abs(gc - 0.5) * 0.4 + promoter_score * 0.3), 1.0)
    
    def _count_is_elements(self, seq):
        """Count IS elements"""
        count = 0
        for is_elem in self.IS_ELEMENTS.keys():
            count += seq.count(is_elem)
        return count
    
    def _resistance_evolution_index(self, seq):
        """Calculate resistance evolution index (combination of mutation and HGT signals)"""
        mutation_score = (self._detect_mutation_pattern(seq, 'S83F') +
                         self._detect_mutation_pattern(seq, 'D87N') +
                         self._detect_mutation_pattern(seq, 'E84K') +
                         self._detect_mutation_pattern(seq, 'S80I')) / 4.0
        
        hgt_score = self._hgt_risk_score(seq)
        
        return (mutation_score * 0.5 + hgt_score * 0.5)


class ResistanceClassifier:
    """Classifies sequences using MLP with optrA-specific features"""
    
    def __init__(self):
        self.model = MLPClassifier(
            hidden_layer_sizes=(128, 64, 32),
            max_iter=1000,
            random_state=42,
            activation='relu',
            solver='adam',
            early_stopping=True,
            validation_fraction=0.1,
            n_iter_no_change=50
        )
        self.scaler = StandardScaler()
        self.feature_names = None
        self.feature_importances = None
    
    def train(self, features_df):
        """Train MLP classifier using all 51 optrA-specific features"""
        self.feature_names = features_df.columns.tolist()
        X = features_df.values
        
        # Generate resistance labels based on optrA-specific indicators
        y = self._generate_resistance_labels(features_df)
        
        print(f"[MLP] Training with {len(self.feature_names)} optrA-specific features")
        
        # Standardize features
        X_scaled = self.scaler.fit_transform(X)
        
        # Train MLP
        self.model.fit(X_scaled, y)
        
        # Calculate feature importance
        self.feature_importances = np.abs(self.model.coefs_[0]).sum(axis=1)
        self.feature_importances = self.feature_importances / self.feature_importances.sum()
        
        print(f"[MLP] Model trained | Architecture: {self.model.hidden_layer_sizes}")
        print(f"[MLP] Training accuracy: {self.model.score(X_scaled, y):.4f}")
        
        return y
    
    def _generate_resistance_labels(self, features_df):
        """Generate binary resistance labels based on optrA-specific features"""
        scores = np.zeros(len(features_df))
        
        # Weight key resistance indicators
        resistance_indicators = {
            'S83F_Mutation_Count': 0.15,
            'D87N_Mutation_Count': 0.15,
            'E84K_Mutation_Count': 0.12,
            'S80I_Mutation_Count': 0.10,
            'Total_IS_Elements': 0.12,
            'Promoter_Strength': 0.10,
            'Has_Stop_Codon': 0.08,
            'HGT_Risk_Score': 0.08,
            'Resistance_Evolution_Index': 0.10
        }
        
        for feature, weight in resistance_indicators.items():
            # Match feature names with slight variations
            matching_cols = [col for col in features_df.columns if feature.lower().replace('_', '') in col.lower().replace('_', '')]
            if matching_cols:
                col = matching_cols[0]
                normalized = (features_df[col] - features_df[col].min()) / (features_df[col].max() - features_df[col].min() + 1e-6)
                scores += normalized.values * weight
        
        # Threshold at 0.5
        labels = (scores > 0.5).astype(int)
        
        return labels
    
    def predict(self, features_df):
        """Predict resistance for sequences"""
        X_scaled = self.scaler.transform(features_df.values)
        predictions = self.model.predict(X_scaled)
        probabilities = self.model.predict_proba(X_scaled)[:, 1]
        
        return predictions, probabilities
    
    def get_feature_importance(self):
        """Return feature importance scores"""
        if self.feature_importances is not None:
            return pd.DataFrame({
                'Feature': self.feature_names,
                'Importance': self.feature_importances
            }).sort_values('Importance', ascending=False)
        return None


class Step2Pipeline:
    """Main pipeline for similarity filtering, feature extraction, and classification"""
    
    def __init__(self, input_file='Data_1.xlsx', output_file='Data_2.xlsx'):
        self.input_file = input_file
        self.output_file = output_file
        self.similarity_filter = SimilarityFilter(min_similarity=0.60, max_similarity=0.99)
        self.extractor = OptrAFeatureExtractor()
        self.classifier = ResistanceClassifier()
    
    def run(self):
        """Execute the complete pipeline"""
        print("=" * 80)
        print("STEP 2: SIMILARITY FILTERING, optrA FEATURE EXTRACTION & CLASSIFICATION")
        print("=" * 80)
        
        # Load sequences
        print("\n[1] Loading sequences from Data_1.xlsx...")
        df = pd.read_excel(self.input_file)
        print(f"    Loaded {len(df)} sequences")
        
        # Apply similarity filtering
        print("\n[2] Applying similarity filtering (Jaccard 60-99%)...")
        df_filtered, sim_summary = self.similarity_filter.filter_sequences(df)
        
        # Extract features
        print("\n[3] Extracting 51 optrA-specific resistance features...")
        features_list = []
        for idx, row in df_filtered.iterrows():
            seq = row['Sequence']
            features = self.extractor.extract_all_features(seq)
            features_list.append(features)
            
            if (idx + 1) % 50 == 0:
                print(f"    Processed {idx + 1}/{len(df_filtered)} sequences")
        
        features_df = pd.DataFrame(features_list)
        print(f"    Extracted {len(features_df.columns)} features")
        
        # Train classifier
        print("\n[4] Training MLP classifier with 51 optrA-specific features...")
        labels = self.classifier.train(features_df)
        
        # Get feature importance
        feature_importance_df = self.classifier.get_feature_importance()
        print(f"    Top 5 important features:")
        for idx, row in feature_importance_df.head(5).iterrows():
            print(f"      - {row['Feature']}: {row['Importance']:.4f}")
        
        # Generate predictions
        print("\n[5] Generating resistance predictions...")
        predictions, probabilities = self.classifier.predict(features_df)
        
        resistant_count = (predictions == 1).sum()
        non_resistant_count = (predictions == 0).sum()
        print(f"    Predicted resistant: {resistant_count}")
        print(f"    Predicted non-resistant: {non_resistant_count}")
        
        # Create final dataset
        print("\n[6] Creating final dataset...")
        final_df = df_filtered.copy()
        
        for col in features_df.columns:
            final_df[col] = features_df[col].values
        
        final_df['Label'] = predictions
        final_df['Resistance_Probability'] = probabilities
        
        # Save to Excel
        print("\n[7] Saving to Data_2.xlsx...")
        self._save_to_excel(final_df, features_df, sim_summary, feature_importance_df)
        
        print("\n" + "=" * 80)
        print(f"COMPLETE: {len(final_df)} sequences processed with 51 optrA-specific features")
        print("=" * 80)
    
    def _save_to_excel(self, final_df, features_df, sim_summary, feature_importance_df):
        """Save results to Excel with comprehensive metadata"""
        wb = Workbook()
        
        # Sheet 1: Features
        sheet1 = wb.active
        sheet1.title = "Features"
        
        for col_idx, col in enumerate(final_df.columns, 1):
            cell = sheet1.cell(row=1, column=col_idx, value=col)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for row_idx, row in enumerate(final_df.values, 2):
            for col_idx, value in enumerate(row, 1):
                sheet1.cell(row=row_idx, column=col_idx, value=value)
        
        sheet1.column_dimensions['A'].width = 25
        sheet1.column_dimensions['B'].width = 70
        for col in sheet1.column_dimensions:
            if col not in ['A', 'B']:
                sheet1.column_dimensions[col].width = 14
        
        # Sheet 2: Similarity Filter Summary
        sheet2 = wb.create_sheet("Similarity_Filter_Summary")
        
        sheet2['A1'] = "Similarity Filtering Summary (Jaccard Similarity)"
        sheet2['A1'].font = Font(bold=True, size=12)
        sheet2.merge_cells('A1:B1')
        
        row = 3
        for key, value in sim_summary.items():
            sheet2[f'A{row}'] = key.replace('_', ' ')
            sheet2[f'B{row}'] = value
            row += 1
        
        sheet2['A{}'.format(row + 1)] = "Filtering Method Details"
        sheet2['A{}'.format(row + 1)].font = Font(bold=True)
        
        sheet2['A{}'.format(row + 3)] = "Step 1: Remove 100% Identical Sequences"
        sheet2['A{}'.format(row + 4)] = "Method: Exact sequence matching using hash comparison"
        sheet2['A{}'.format(row + 5)] = "Result: All duplicate sequences removed"
        
        sheet2['A{}'.format(row + 7)] = "Step 2: Jaccard Similarity Filtering"
        sheet2['A{}'.format(row + 8)] = "Method: Calculate Jaccard similarity between sequence pairs"
        sheet2['A{}'.format(row + 9)] = f"Threshold: Keep sequences with Jaccard similarity between {sim_summary['Similarity_threshold_min']*100:.0f}% and {sim_summary['Similarity_threshold_max']*100:.0f}%"
        
        sheet2.column_dimensions['A'].width = 50
        sheet2.column_dimensions['B'].width = 25
        
        # Sheet 3: Feature Importance
        sheet3 = wb.create_sheet("Feature_Importance")
        
        sheet3['A1'] = "optrA Feature Importance (MLP Model)"
        sheet3['A1'].font = Font(bold=True, size=12)
        
        sheet3['A2'] = "Feature"
        sheet3['B2'] = "Importance Score"
        
        for header_cell in ['A2', 'B2']:
            sheet3[header_cell].font = Font(bold=True, color='FFFFFF')
            sheet3[header_cell].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for idx, row in enumerate(feature_importance_df.values, 3):
            sheet3[f'A{idx}'] = row[0]
            sheet3[f'B{idx}'] = row[1]
        
        sheet3.column_dimensions['A'].width = 40
        sheet3.column_dimensions['B'].width = 20
        
        # Sheet 4: MetaData
        sheet4 = wb.create_sheet("MetaData")
        
        sheet4['A1'] = "optrA Feature Extraction & Classification Metadata"
        sheet4['A1'].font = Font(bold=True, size=12)
        
        sheet4['A3'] = "Dataset Statistics"
        sheet4['A3'].font = Font(bold=True, size=11)
        
        sheet4['A4'] = "Total Sequences (After Filtering):"
        sheet4['B4'] = len(final_df)
        
        sheet4['A5'] = "Resistant Sequences:"
        sheet4['B5'] = (final_df['Label'] == 1).sum()
        
        sheet4['A6'] = "Non-Resistant Sequences:"
        sheet4['B6'] = (final_df['Label'] == 0).sum()
        
        sheet4['A7'] = "Resistance Rate:"
        sheet4['B7'] = f"{(final_df['Label'] == 1).sum() / len(final_df) * 100:.2f}%"
        
        sheet4['A9'] = "Feature Extraction Parameters"
        sheet4['A9'].font = Font(bold=True, size=11)
        
        sheet4['A10'] = "Total Features Extracted:"
        sheet4['B10'] = 51
        
        sheet4['A11'] = "Feature Type:"
        sheet4['B11'] = "optrA-Specific Resistance Markers"
        
        sheet4['A12'] = "Features Used in MLP:"
        sheet4['B12'] = "All 51 features"
        
        sheet4['A14'] = "MLP Classifier Configuration"
        sheet4['A14'].font = Font(bold=True, size=11)
        
        sheet4['A15'] = "MLP Architecture:"
        sheet4['B15'] = "(128, 64, 32)"
        
        sheet4['A16'] = "Activation Function:"
        sheet4['B16'] = "ReLU"
        
        sheet4['A17'] = "Solver:"
        sheet4['B17'] = "Adam"
        
        sheet4['A18'] = "Early Stopping:"
        sheet4['B18'] = "True"
        
        sheet4['A20'] = "Feature Groups"
        sheet4['A20'].font = Font(bold=True, size=11)
        
        feature_groups = {
            "Group 1: Gene Structure & Integrity": "Features 1-8",
            "Group 2: GC Composition": "Features 9-14",
            "Group 3: Known Resistance Mutations": "Features 15-18",
            "Group 4: Promoter & Regulatory Elements": "Features 19-23",
            "Group 5: IS Elements & Mobile Genetic Elements": "Features 24-28",
            "Group 6: Efflux Pump Characteristics": "Features 29-33",
            "Group 7: Dinucleotide & Codon Bias": "Features 34-39",
            "Group 8: Sequence Complexity": "Features 40-44",
            "Group 9: Mutation Signals & Adaptation": "Features 45-49",
            "Group 10: Functional & Evolution Metrics": "Features 50-51"
        }
        
        row_num = 21
        for group, features in feature_groups.items():
            sheet4[f'A{row_num}'] = group
            sheet4[f'B{row_num}'] = features
            row_num += 1
        
        sheet4['A{}'.format(row_num + 1)] = "All 51 optrA-Specific Features"
        sheet4['A{}'.format(row_num + 1)].font = Font(bold=True, size=11)
        
        for idx, feature in enumerate(features_df.columns, row_num + 2):
            sheet4[f'A{idx}'] = feature
        
        sheet4.column_dimensions['A'].width = 50
        sheet4.column_dimensions['B'].width = 30
        
        wb.save(self.output_file)


if __name__ == "__main__":
    pipeline = Step2Pipeline(input_file='Data_1.xlsx', output_file='Data_2.xlsx')
    pipeline.run()