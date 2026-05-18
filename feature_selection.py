"""
Step 4: Feature Selection & Data Splitting (ALL FEATURES AT EACH STEP)
Reads Distill_RoBERTa_Embeddings sheet from Data_3.xlsx
Input features = 51 optrA features  +  768 Distill-RoBERTa embeddings (819 total)
1. Pearson correlation filtering on ALL 819 features (remove only high correlations ≥ 0.90)
2. PCA on ALL filtered features (85% variance - keeps all needed components)
3. Feature Selection: Boruta, SVM-RFE, Filter-Based on ALL PCA components
4. Select best method (most features retained)
5. Chi-Square analysis on train/val/test splits
6. Apply best split with Type column (TRAIN/VALIDATE/TEST)
Input:  Data_3.xlsx  (Distill_RoBERTa_Embeddings sheet)
Output: Data_4.xlsx
"""

# Imported modules
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier
from scipy.stats import pearsonr, chi2_contingency
import warnings

warnings.filterwarnings('ignore')

try:
    from boruta import BorutaPy
    BORUTA_AVAILABLE = True
except ImportError:
    BORUTA_AVAILABLE = False
    print("⚠️ BorutaPy not installed. Using RandomForest feature importance instead.")

# ========================== UTILITY FUNCTIONS ==========================

class ColumnIdentifier:
    """Identify column types in dataframe"""

    # Columns that are never treated as features
    METADATA_COLS   = ['Header', 'Sequence', 'Sequence_Length', 'Source']
    LABEL_COLS      = ['Label', 'Resistance_Probability']
    SPLIT_TYPE_COL  = 'Type'

    @staticmethod
    def get_columns(df):
        """
        Returns a dict with keys:
          metadata   – identity / sequence columns  (kept as-is in every output sheet)
          labels     – Label + Resistance_Probability  (re-attached to final sheet only)
          features   – numbered optrA feature columns  (e.g. 1_Gene_Length_bp … 51_…)
          embeddings – Embedding_RoBERTa_* columns
        """
        non_feature = set(ColumnIdentifier.METADATA_COLS +
                          ColumnIdentifier.LABEL_COLS +
                          [ColumnIdentifier.SPLIT_TYPE_COL])

        metadata_cols   = [c for c in ColumnIdentifier.METADATA_COLS if c in df.columns]
        label_cols      = [c for c in ColumnIdentifier.LABEL_COLS     if c in df.columns]
        embedding_cols  = [c for c in df.columns if c.startswith('Embedding_')]
        feature_cols    = [c for c in df.columns
                           if c not in non_feature and c not in embedding_cols]

        return {
            'metadata':   metadata_cols,
            'labels':     label_cols,
            'features':   feature_cols,
            'embeddings': embedding_cols,
        }

# ========================== STEP 1: PEARSON FILTERING ==========================

class PearsonFeatureFilter:
    """Filter features by Pearson correlation – REMOVE ONLY HIGH CORRELATIONS"""

    def __init__(self, correlation_threshold=0.90):
        self.correlation_threshold = correlation_threshold
        self.removed_count  = 0
        self.retained_count = 0

    def filter_features(self, X, y, feature_names):
        """
        Remove one member of each highly correlated pair (> threshold).
        When both are equally correlated with y, keep the first index.
        """
        print(f"\n{'='*80}")
        print(f"[STEP 1] PEARSON CORRELATION FILTERING")
        print(f"{'='*80}")
        print(f"[Pearson] Threshold: {self.correlation_threshold}")
        print(f"[Pearson] Input features: {X.shape[1]}")

        X = X.astype(float)
        y = y.astype(float)

        to_remove = set()

        for i in range(X.shape[1]):
            if i in to_remove:
                continue
            for j in range(i + 1, X.shape[1]):
                if j in to_remove:
                    continue
                try:
                    corr = abs(pearsonr(X[:, i], X[:, j])[0])
                    if corr > self.correlation_threshold:
                        target_corr_i = abs(pearsonr(X[:, i], y)[0]) if np.std(X[:, i]) > 1e-8 else 0
                        target_corr_j = abs(pearsonr(X[:, j], y)[0]) if np.std(X[:, j]) > 1e-8 else 0
                        # Drop whichever is less correlated with the label
                        to_remove.add(i if target_corr_i < target_corr_j else j)
                except Exception:
                    pass

        selected_indices = [i for i in range(X.shape[1]) if i not in to_remove]
        self.removed_count  = len(to_remove)
        self.retained_count = len(selected_indices)

        print(f"[Pearson] Removed : {self.removed_count}")
        print(f"[Pearson] Retained: {self.retained_count}\n")

        if not selected_indices:
            print("⚠️  No features retained after Pearson filter – reverting to all features.")
            return X, np.arange(X.shape[1]), feature_names

        idx = np.array(selected_indices)
        return X[:, idx], idx, feature_names[idx]

# ========================== STEP 2: PCA ==========================

class PCAFeatureSelector:
    """Apply PCA – KEEP ALL COMPONENTS needed for 85 % variance"""

    def __init__(self, variance_threshold=0.85):
        self.variance_threshold = variance_threshold
        self.pca            = None
        self.removed_count  = 0
        self.retained_count = 0

    def fit_transform(self, X):
        print(f"{'='*80}")
        print(f"[STEP 2] PCA DIMENSIONALITY REDUCTION")
        print(f"{'='*80}")
        print(f"[PCA] Target variance: {self.variance_threshold*100:.0f}%")
        print(f"[PCA] Input features : {X.shape[1]}")

        X = X.astype(float)
        scaler  = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        pca   = PCA(n_components=self.variance_threshold, random_state=42)
        X_pca = pca.fit_transform(X_scaled)

        self.pca            = pca
        self.removed_count  = X.shape[1] - pca.n_components_
        self.retained_count = pca.n_components_

        explained = np.sum(pca.explained_variance_ratio_) * 100
        print(f"[PCA] Explained variance : {explained:.2f}%")
        print(f"[PCA] Components retained: {self.retained_count}")
        print(f"[PCA] Removed            : {self.removed_count}\n")

        pca_names = np.array([f'PCA_Comp_{i+1}' for i in range(pca.n_components_)])
        return X_pca, np.arange(pca.n_components_), pca_names

# ========================== STEP 3A: BORUTA ==========================

class BorutaSelector:
    """Boruta – SELECT ALL confirmed features"""

    def __init__(self):
        self.selected_count = 0

    def select_features(self, X, y):
        print(f"{'='*80}")
        print(f"[STEP 3A] FEATURE SELECTION – BORUTA")
        print(f"{'='*80}")
        print(f"[Boruta] Input features: {X.shape[1]}")

        X = X.astype(float)
        y = y.astype(int)

        scaler   = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        if BORUTA_AVAILABLE:
            rf     = RandomForestClassifier(n_jobs=-1, random_state=42, n_estimators=100)
            boruta = BorutaPy(rf, n_estimators='auto', random_state=42, verbose=0)
            boruta.fit(X_scaled, y)
            selected_idx = np.where(boruta.support_)[0]
        else:
            # Fallback: keep features above mean importance
            rf = RandomForestClassifier(n_jobs=-1, random_state=42, n_estimators=100)
            rf.fit(X_scaled, y)
            importances  = rf.feature_importances_
            selected_idx = np.where(importances > np.mean(importances))[0]

        self.selected_count = len(selected_idx)
        print(f"[Boruta] Selected: {self.selected_count}")
        print(f"[Boruta] Removed : {X.shape[1] - self.selected_count}\n")
        return X[:, selected_idx], selected_idx

# ========================== STEP 3B: SVM-RFE ==========================

class SVMRFESelector:
    """SVM-RFE – RANK and KEEP ALL features"""

    def __init__(self):
        self.selected_count = 0

    def select_features(self, X, y):
        print(f"{'='*80}")
        print(f"[STEP 3B] FEATURE SELECTION – SVM-RFE")
        print(f"{'='*80}")
        print(f"[SVM-RFE] Input features: {X.shape[1]}")

        X = X.astype(float)
        y = y.astype(int)

        scaler   = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        svm         = SVC(kernel='linear', random_state=42, max_iter=10000)
        svm.fit(X_scaled, y)
        importances = np.abs(svm.coef_[0])
        selected_idx = np.argsort(importances)   # keep all, sorted ascending by importance

        self.selected_count = len(selected_idx)
        print(f"[SVM-RFE] Selected: {self.selected_count} (ALL, sorted by importance)")
        print(f"[SVM-RFE] Removed : 0\n")
        return X[:, selected_idx], selected_idx

# ========================== STEP 3C: FILTER-BASED ==========================

class FilterBasedSelector:
    """Filter-Based (F-statistic) – RANK and KEEP ALL features"""

    def __init__(self):
        self.selected_count = 0

    def select_features(self, X, y):
        print(f"{'='*80}")
        print(f"[STEP 3C] FEATURE SELECTION – FILTER-BASED")
        print(f"{'='*80}")
        print(f"[Filter-Based] Input features: {X.shape[1]}")

        X = X.astype(float)
        y = y.astype(int)

        from sklearn.feature_selection import f_classif
        scores, _    = f_classif(X, y)
        selected_idx = np.argsort(scores)[::-1]   # descending F-score order

        self.selected_count = len(selected_idx)
        print(f"[Filter-Based] Selected: {self.selected_count} (ALL, sorted by F-score)")
        print(f"[Filter-Based] Removed : 0\n")
        return X[:, selected_idx], selected_idx

# ========================== CHI-SQUARE ANALYSIS ==========================

class ChiSquareAnalyzer:
    """Analyse different train/val/test split ratios with a chi-square test"""

    def analyze_splits(self, y, splits):
        print(f"\n{'='*80}")
        print(f"[CHI-SQUARE] Analysing {len(splits)} split configurations...")
        print(f"{'='*80}")

        results = []
        for tr, va, te in splits:
            np.random.seed(42)
            idx      = np.random.permutation(len(y))
            tr_size  = int(len(y) * tr / 100)
            va_size  = int(len(y) * va / 100)

            tr_labels = y[idx[:tr_size]]
            va_labels = y[idx[tr_size:tr_size + va_size]]
            te_labels = y[idx[tr_size + va_size:]]

            contingency = np.array([
                [np.sum(tr_labels == 0), np.sum(tr_labels == 1)],
                [np.sum(va_labels == 0), np.sum(va_labels == 1)],
                [np.sum(te_labels == 0), np.sum(te_labels == 1)],
            ])

            try:
                chi2, pval, _, _ = chi2_contingency(contingency)
            except Exception:
                chi2, pval = 0.0, 1.0

            balance = min(tr, va, te) / max(tr, va, te)
            results.append({
                'Split_Ratio':       f'{tr}-{va}-{te}',
                'Train_%':           tr,
                'Validate_%':        va,
                'Test_%':            te,
                'Train_Samples':     tr_size,
                'Validate_Samples':  va_size,
                'Test_Samples':      len(y) - tr_size - va_size,
                'Chi2_Statistic':    chi2,
                'P_Value':           pval,
                'Balance_Score':     balance,
            })
            print(f"  {tr}-{va}-{te}: Train={tr_size}, Val={va_size}, "
                  f"Test={len(y)-tr_size-va_size}, Balance={balance:.4f}")

        return pd.DataFrame(results)

    def find_best_split(self, chi_results_df):
        best_idx = chi_results_df['Balance_Score'].idxmax()
        return chi_results_df.iloc[best_idx], best_idx

# ========================== DATA SPLITTING ==========================

class DataSplitter:
    """Split data into TRAIN / VALIDATE / TEST and add a 'Type' column"""

    @staticmethod
    def apply_split(df, train_pct, val_pct):
        print(f"\n{'='*80}")
        print(f"[DATA SPLIT] Applying {train_pct}-{val_pct}-{100-train_pct-val_pct} split...")
        print(f"{'='*80}")

        df_shuffled = df.sample(frac=1, random_state=42).reset_index(drop=True)
        n       = len(df_shuffled)
        n_train = int(n * train_pct / 100)
        n_val   = int(n * val_pct   / 100)

        df_train = df_shuffled[:n_train].copy();            df_train['Type'] = 'TRAIN'
        df_val   = df_shuffled[n_train:n_train+n_val].copy(); df_val['Type'] = 'VALIDATE'
        df_test  = df_shuffled[n_train+n_val:].copy();      df_test['Type']  = 'TEST'

        result = pd.concat([df_train, df_val, df_test], ignore_index=True)
        print(f"  TRAIN:    {len(df_train)}")
        print(f"  VALIDATE: {len(df_val)}")
        print(f"  TEST:     {len(df_test)}\n")
        return result

# ========================== EXCEL WRITER ==========================

class ExcelWriter:

    @staticmethod
    def write(output_file, df_pearson, df_pca, df_final_selected,
              chi2_results, best_idx, best_method, pipeline_stats):
        print(f"{'='*80}")
        print(f"[EXCEL] Writing to {output_file} ...")
        print(f"{'='*80}\n")

        wb = Workbook()
        wb.remove(wb.active)

        ExcelWriter._add_df_sheet(wb, "Pearson_Filtered",      df_pearson)
        ExcelWriter._add_df_sheet(wb, "PCA_Filtered",          df_pca)
        ExcelWriter._add_df_sheet(wb, "Filter_Based_Filtered", df_final_selected)

        # ── Analytics sheet ──────────────────────────────────────────────────
        ws = wb.create_sheet("Analytics")
        ws['A1'] = "Feature Selection & Data Splitting Analytics"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = "Model: Distill_RoBERTa (F1=0.6636, Accuracy=0.6907)"
        ws['A3'] = f"Input features: 51 optrA features + 768 Distill-RoBERTa embeddings = 819 total"
        ws['A4'] = f"Best Selection Method: {best_method}"

        ws['A6'] = "Pipeline Summary"
        ws['A6'].font = Font(bold=True, size=11)

        hdr_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF')
        for col_idx, hdr in enumerate(['Stage', 'Features_Removed', 'Features_Retained', 'Notes'], 1):
            c = ws.cell(7, col_idx, hdr)
            c.font = hdr_font
            c.fill = hdr_fill

        for row_idx, stat in enumerate(pipeline_stats, 8):
            ws.cell(row_idx, 1, stat['Stage'])
            ws.cell(row_idx, 2, stat['Removed'])
            ws.cell(row_idx, 3, stat['Retained'])
            ws.cell(row_idx, 4, stat.get('Method', stat.get('Threshold', '')))

        for col in ('A', 'B', 'C', 'D'):
            ws.column_dimensions[col].width = 30

        # ── Chi-Square Analytics sheet ────────────────────────────────────────
        ws = wb.create_sheet("Chi_Square_Analytics")
        ws['A1'] = "Chi-Square Analysis of Train / Validate / Test Splits"
        ws['A1'].font = Font(bold=True, size=12)
        ExcelWriter._write_dataframe(ws, chi2_results, start_row=3)

        # Highlight best split row in yellow
        if best_idx >= 0:
            best_row = best_idx + 4        # header at row 3, data starts at row 4
            yellow   = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for col in range(1, chi2_results.shape[1] + 1):
                ws.cell(best_row, col).fill = yellow

        # ── Data Split sheet ──────────────────────────────────────────────────
        ExcelWriter._add_df_sheet(wb, "Data_Split", df_final_selected)

        wb.save(output_file)
        print(f"✅ Saved: {output_file}")

    # ── helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _add_df_sheet(wb, sheet_name, df):
        ws = wb.create_sheet(sheet_name)
        ExcelWriter._write_dataframe(ws, df)

    @staticmethod
    def _write_dataframe(ws, df, start_row=1):
        hdr_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF')

        for col_idx, col_name in enumerate(df.columns, 1):
            c = ws.cell(start_row, col_idx, col_name)
            c.font = hdr_font
            c.fill = hdr_fill

        for row_idx, row_data in enumerate(df.values, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                c = ws.cell(row_idx, col_idx, value)
                if isinstance(value, float):
                    c.number_format = '0.0000'

        for col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].width = 12
        ws.column_dimensions['A'].width = 25
        if 'Sequence' in df.columns:
            seq_col = df.columns.get_loc('Sequence') + 1
            ws.column_dimensions[
                ws.cell(1, seq_col).column_letter
            ].width = 70

# ========================== MAIN PIPELINE ==========================

class FeatureSelectionPipeline:
    """
    Full feature-selection and data-splitting pipeline.

    KEY CHANGE vs. original:
        self.X  now contains  51 optrA features  +  768 Distill-RoBERTa embeddings
        (total 819 columns).  All selection steps operate on this combined matrix.
    """

    def __init__(self,
                 input_file='Data_3.xlsx',
                 output_file='Data_4.xlsx'):
        self.input_file     = input_file
        self.output_file    = output_file
        self.df_original    = None
        self.X              = None
        self.y              = None
        self.feature_names  = None
        self.pipeline_stats = []

    # ── public entry point ────────────────────────────────────────────────────

    def run(self):
        print("\n" + "="*80)
        print("STEP 4: FEATURE SELECTION & DATA SPLITTING")
        print("MODEL: Distill_RoBERTa (F1=0.6636, Accuracy=0.6907)")
        print("INPUT: 51 optrA features + 768 Distill-RoBERTa embeddings = 819 features")
        print("="*80)

        # [0] Load
        self._load_data()

        # [1] Pearson filter
        pearson = PearsonFeatureFilter(correlation_threshold=0.90)
        X_pearson, _, names_pearson = pearson.filter_features(
            self.X, self.y, self.feature_names)
        df_pearson = self._build_output_df(X_pearson, names_pearson)
        self.pipeline_stats.append({
            'Stage':    'Pearson_Correlation',
            'Removed':  pearson.removed_count,
            'Retained': pearson.retained_count,
            'Threshold': '0.90',
        })

        # [2] PCA
        pca = PCAFeatureSelector(variance_threshold=0.85)
        X_pca, _, names_pca = pca.fit_transform(X_pearson)
        df_pca = self._build_output_df(X_pca, names_pca)
        self.pipeline_stats.append({
            'Stage':    'PCA_85_Variance',
            'Removed':  pca.removed_count,
            'Retained': pca.retained_count,
            'Threshold': '85% variance',
        })

        # [3] Three selection methods (all keep all PCA components)
        boruta     = BorutaSelector()
        X_boruta, _ = boruta.select_features(X_pca, self.y)

        svm_rfe    = SVMRFESelector()
        X_svm, _   = svm_rfe.select_features(X_pca, self.y)

        filter_sel = FilterBasedSelector()
        X_filter, _ = filter_sel.select_features(X_pca, self.y)

        # [4] Pick method that retained the most features
        methods = {
            'Boruta':       (X_boruta.shape[1],  X_boruta,  boruta.selected_count),
            'SVM-RFE':      (X_svm.shape[1],     X_svm,     svm_rfe.selected_count),
            'Filter-Based': (X_filter.shape[1],  X_filter,  filter_sel.selected_count),
        }
        best_method = max(methods, key=lambda k: methods[k][0])
        best_n, X_final, best_count = methods[best_method]

        print(f"\n{'='*80}")
        print(f"✅ Best Method: {best_method}  ({best_count} features retained)")
        print(f"{'='*80}\n")

        self.pipeline_stats.append({
            'Stage':    f'{best_method}_Selected',
            'Removed':  X_pca.shape[1] - best_count,
            'Retained': best_count,
            'Method':   best_method,
        })

        feature_col_names = [f'Feature_{i}' for i in range(X_final.shape[1])]
        df_final = self._build_output_df(X_final, feature_col_names)

        # [5] Chi-Square split analysis
        splits = [(70, 15, 15), (65, 20, 15), (75, 15, 10), (60, 20, 20), (80, 10, 10)]
        chi_analyzer = ChiSquareAnalyzer()
        chi_results  = chi_analyzer.analyze_splits(self.y, splits)
        best_split, best_idx = chi_analyzer.find_best_split(chi_results)

        print(f"{'='*80}")
        print(f"✅ Best Split: {best_split['Split_Ratio']} "
              f"(Balance Score: {best_split['Balance_Score']:.4f})")
        print(f"{'='*80}\n")

        # [6] Apply split (re-attach Label before splitting)
        df_final_with_label = df_final.copy()
        df_final_with_label['Label'] = self.y
        df_with_type = DataSplitter.apply_split(
            df_final_with_label,
            int(best_split['Train_%']),
            int(best_split['Validate_%']),
        )

        # [7] Write Excel
        ExcelWriter.write(
            self.output_file,
            df_pearson, df_pca, df_with_type,
            chi_results, best_idx, best_method,
            self.pipeline_stats,
        )

        print(f"{'='*80}")
        print("✅ COMPLETE: Step 4 finished successfully!")
        print(f"{'='*80}\n")

    # ── private helpers ───────────────────────────────────────────────────────

    def _load_data(self):
        """
        Load the Distill_RoBERTa_Embeddings sheet.

        self.X           = combined matrix of 51 optrA features + 768 embeddings
        self.y           = Label column (0/1)
        self.feature_names = column names for self.X
        """
        print(f"\n{'='*80}")
        print(f"[STEP 0] LOADING DATA")
        print(f"{'='*80}")
        print(f"File : {self.input_file}")
        print(f"Sheet: Distill_RoBERTa_Embeddings")

        self.df_original = pd.read_excel(
            self.input_file,
            sheet_name='Distill_RoBERTa_Embeddings',
        )

        cols = ColumnIdentifier.get_columns(self.df_original)

        feature_cols   = cols['features']    # 51 optrA columns
        embedding_cols = cols['embeddings']  # 768 Embedding_RoBERTa_* columns

        all_input_cols = feature_cols + embedding_cols

        print(f"\n✅ Data loaded:")
        print(f"   Rows              : {len(self.df_original)}")
        print(f"   Total columns     : {len(self.df_original.columns)}")
        print(f"   optrA features    : {len(feature_cols)}")
        print(f"   RoBERTa embeddings: {len(embedding_cols)}")
        print(f"   Combined X cols   : {len(all_input_cols)}")

        self.X            = self.df_original[all_input_cols].astype(float).values
        self.y            = self.df_original['Label'].astype(int).values
        self.feature_names = np.array(all_input_cols)

        unique, counts = np.unique(self.y, return_counts=True)
        print(f"\n   Class Distribution:")
        for u, c in zip(unique, counts):
            label_name = "Resistant" if u == 1 else "Non-Resistant"
            print(f"     {label_name} ({u}): {c}  ({c/len(self.y)*100:.2f}%)")

    def _build_output_df(self, X, feature_names):
        """
        Build an output DataFrame:
          [metadata cols] + [feature / PCA columns]

        Note: because embeddings are now part of the feature matrix X (and transformed / filtered through Pearson / PCA),
        we do NOT re-append the raw embedding columns from df_original here.
        """
        cols = ColumnIdentifier.get_columns(self.df_original)

        df_features = pd.DataFrame(X, columns=feature_names)

        # Prepend metadata (Header, Sequence, Sequence_Length, Source)
        for col in reversed(cols['metadata']):
            if col in self.df_original.columns:
                df_features.insert(0, col, self.df_original[col].values)

        return df_features

# ====================== ENTRY POINT ======================

if __name__ == "__main__":
    pipeline = FeatureSelectionPipeline(
        input_file='/kaggle/input/datasets/giogogi/data-3/Data_3.xlsx',
        output_file='Data_4.xlsx',
    )
    pipeline.run()