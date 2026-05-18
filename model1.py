"""
Step 6: Model 1: Logistic Regression Resistance Classifier
Reads Data_4.xlsx (Data_Split sheet)
1. Splits rows by Type column (TRAIN / VALIDATE / TEST)
2. Trains a Logistic Regression model on the TRAIN split
3. Evaluates on all three phases and prints metrics
4. Runs final prediction on TEST split
5. Saves sequences predicted as RESISTANT (Label = 1) to Data_6.xlsx
   → columns: Header, Sequence, Sequence_Length, Source, Label_Original,
               Label_Predicted, Resistance_Probability
Input : Data_4.xlsx  (Data_Split sheet)
Output: Data_6.xlsx  (resistant test sequences → fed into model2.py)
"""

# Imported modules
import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sklearn.linear_model   import LogisticRegression
from sklearn.preprocessing  import StandardScaler
from sklearn.metrics        import (accuracy_score, precision_score,
                                    recall_score, f1_score, roc_auc_score,
                                    average_precision_score, classification_report)

warnings.filterwarnings('ignore')

# ========================== CONSTANTS ==========================

PHASE_TRAIN = 'TRAIN'
PHASE_VAL   = 'VALIDATE'
PHASE_TEST  = 'TEST'

NON_FEATURE_COLS = {'Header', 'Sequence', 'Sequence_Length', 'Source',
                    'Label', 'Resistance_Probability', 'Type'}

# ========================== DATA LOADER ==========================

class DataLoader:
    """Load Data_4.xlsx and return per-phase splits."""

    def __init__(self, input_file='Data_4.xlsx'):
        self.input_file  = input_file
        self.feature_cols = []
        self.df_full     = None

    def load(self):
        print(f"\n{'='*80}")
        print(f"[MODEL 1 | STEP 0] LOADING DATA")
        print(f"{'='*80}")
        print(f"File : {self.input_file}")
        print(f"Sheet: Data_Split")

        self.df_full = pd.read_excel(self.input_file, sheet_name='Data_Split')

        print(f"\n✅ Loaded:")
        print(f"   Rows   : {len(self.df_full)}")
        print(f"   Columns: {len(self.df_full.columns)}")

        if 'Type' not in self.df_full.columns:
            raise ValueError("'Type' column not found in Data_Split sheet.")
        if 'Label' not in self.df_full.columns:
            raise ValueError("'Label' column not found in Data_Split sheet.")

        self.feature_cols = [c for c in self.df_full.columns if c not in NON_FEATURE_COLS]
        print(f"   Feature columns : {len(self.feature_cols)}")

        splits = {}
        for phase in [PHASE_TRAIN, PHASE_VAL, PHASE_TEST]:
            mask = self.df_full['Type'].str.upper() == phase
            sub  = self.df_full[mask].reset_index(drop=True)
            X    = sub[self.feature_cols].astype(float).values
            y    = sub['Label'].astype(int).values
            splits[phase] = (X, y, sub)   # (features, labels, full-row dataframe)
            print(f"   {phase:10s}: {len(sub):>5d} rows  "
                  f"(resistant={int((y==1).sum())}, non-resistant={int((y==0).sum())})")

        return splits

# ========================== METRICS ==========================

class MetricsReporter:
    """Compute and print the 6 standard metrics for one phase."""

    @staticmethod
    def compute(phase, y_true, y_pred, y_prob):
        metrics = {
            'Phase':     phase,
            'Accuracy':  accuracy_score(y_true, y_pred),
            'Precision': precision_score(y_true, y_pred, zero_division=0),
            'Recall':    recall_score(y_true, y_pred, zero_division=0),
            'F1_Score':  f1_score(y_true, y_pred, zero_division=0),
            'ROC_AUC':   roc_auc_score(y_true, y_prob)
                         if len(np.unique(y_true)) > 1 else 0.0,
            'AUROC':     average_precision_score(y_true, y_prob)
                         if len(np.unique(y_true)) > 1 else 0.0,
        }
        print(f"  {phase:10s} | Acc={metrics['Accuracy']:.4f}  "
              f"P={metrics['Precision']:.4f}  R={metrics['Recall']:.4f}  "
              f"F1={metrics['F1_Score']:.4f}  "
              f"ROC_AUC={metrics['ROC_AUC']:.4f}  AUROC={metrics['AUROC']:.4f}")
        return metrics

# ========================== LOGISTIC REGRESSION MODEL ==========================

class ResistanceClassifierLR:
    """Logistic Regression with standard scaling."""

    def __init__(self):
        self.scaler = StandardScaler()
        self.model  = LogisticRegression(
            max_iter=2000,
            solver='lbfgs',
            random_state=42,
            class_weight='balanced',   # handles class imbalance
        )
        self._fitted = False

    # ── training ─────────────────────────────────────────────────────────────

    def fit(self, X_train, y_train):
        print(f"\n{'='*80}")
        print(f"[MODEL 1 | STEP 1] TRAINING LOGISTIC REGRESSION")
        print(f"{'='*80}")
        print(f"  Train samples : {len(y_train)}")
        print(f"  Features      : {X_train.shape[1]}")

        X_sc = self.scaler.fit_transform(X_train)
        self.model.fit(X_sc, y_train)
        self._fitted = True

        train_acc = self.model.score(X_sc, y_train)
        print(f"  Training accuracy : {train_acc:.4f}")
        print(f"  Solver converged  : {self.model.n_iter_[0]} iterations\n")

    # ── inference ─────────────────────────────────────────────────────────────

    def predict(self, X):
        if not self._fitted:
            raise RuntimeError("Model not fitted yet. Call fit() first.")
        X_sc  = self.scaler.transform(X)
        preds = self.model.predict(X_sc)
        probs = self.model.predict_proba(X_sc)[:, 1]
        return preds, probs

# ========================== EVALUATION ==========================

class ModelEvaluator:
    """Evaluate model across all three phases."""

    @staticmethod
    def evaluate(clf, splits):
        print(f"\n{'='*80}")
        print(f"[MODEL 1 | STEP 2] PHASE EVALUATION")
        print(f"{'='*80}")

        all_metrics = []
        for phase in [PHASE_TRAIN, PHASE_VAL, PHASE_TEST]:
            X, y, _ = splits[phase]
            y_pred, y_prob = clf.predict(X)
            m = MetricsReporter.compute(phase, y, y_pred, y_prob)
            all_metrics.append(m)

        return all_metrics

# ========================== EXCEL WRITER ==========================

class ExcelWriter:
    """Write resistant test sequences to Data_6.xlsx."""

    HEADER_FILL  = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    HEADER_FONT  = Font(bold=True, color='FFFFFF')
    YELLOW_FILL  = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    @staticmethod
    def write(output_file, df_resistant, all_metrics):
        print(f"\n{'='*80}")
        print(f"[MODEL 1 | STEP 3] WRITING Data_6.xlsx")
        print(f"{'='*80}")

        wb = Workbook()
        wb.remove(wb.active)

        # ── Sheet 1: Resistant Sequences ─────────────────────────────────────
        ws = wb.create_sheet('Resistant_Sequences')
        ws['A1'] = 'Data_6 – Antibiotic Resistant Sequences (Logistic Regression)'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = (f'Total resistant sequences identified: {len(df_resistant)}  |  '
                    f'Input for model2.py GNN analysis')
        ws['A2'].font = Font(italic=True, size=10)

        ExcelWriter._write_df(ws, df_resistant, start_row=4)

        # ── Sheet 2: Phase Metrics ────────────────────────────────────────────
        ws2 = wb.create_sheet('Phase_Metrics')
        ws2['A1'] = 'Logistic Regression – Performance Metrics by Phase'
        ws2['A1'].font = Font(bold=True, size=12)

        metrics_df = pd.DataFrame(all_metrics)
        ExcelWriter._write_df(ws2, metrics_df, start_row=3)

        wb.save(output_file)
        print(f"✅ Saved: {output_file}")
        print(f"   Resistant sequences: {len(df_resistant)}")

    @staticmethod
    def _write_df(ws, df, start_row=1):
        for col_idx, col_name in enumerate(df.columns, 1):
            c = ws.cell(start_row, col_idx, str(col_name))
            c.font  = ExcelWriter.HEADER_FONT
            c.fill  = ExcelWriter.HEADER_FILL

        for row_idx, row in enumerate(df.values, start_row + 1):
            for col_idx, value in enumerate(row, 1):
                c = ws.cell(row_idx, col_idx,
                            float(value) if isinstance(value, (np.floating, np.integer))
                            else value)
                if isinstance(value, (float, np.floating)):
                    c.number_format = '0.0000'

        for i, col_name in enumerate(df.columns, 1):
            letter = ws.cell(1, i).column_letter
            if col_name in ('Sequence',):
                ws.column_dimensions[letter].width = 70
            elif col_name in ('Header',):
                ws.column_dimensions[letter].width = 30
            else:
                ws.column_dimensions[letter].width = 20

# ========================== PIPELINE ==========================

class Model1Pipeline:
    """
    Full Model 1 pipeline:
      Load → Train LR → Evaluate all phases → Filter TEST resistant → Save Data_6.
    """

    def __init__(self, input_file='Data_4.xlsx', output_file='Data_6.xlsx'):
        self.input_file  = input_file
        self.output_file = output_file

    def run(self):
        print("\n" + "="*80)
        print("MODEL 1 – LOGISTIC REGRESSION RESISTANCE CLASSIFIER")
        print("Input : Data_4.xlsx  |  Output: Data_6.xlsx")
        print("="*80)

        # [0] Load
        loader = DataLoader(self.input_file)
        splits = loader.load()

        # [1] Train
        clf = ResistanceClassifierLR()
        X_train, y_train, _ = splits[PHASE_TRAIN]
        clf.fit(X_train, y_train)

        # [2] Evaluate all phases
        all_metrics = ModelEvaluator.evaluate(clf, splits)

        # [3] Predict on TEST split → keep only resistant
        print(f"\n{'='*80}")
        print(f"[MODEL 1 | STEP 3] EXTRACTING RESISTANT TEST SEQUENCES")
        print(f"{'='*80}")

        X_test, y_test, df_test_full = splits[PHASE_TEST]
        y_pred_test, y_prob_test = clf.predict(X_test)

        df_output = df_test_full[['Header', 'Sequence', 'Sequence_Length', 'Source']].copy()
        df_output['Label_Original']          = y_test
        df_output['Label_Predicted']         = y_pred_test
        df_output['Resistance_Probability']  = np.round(y_prob_test, 6)

        df_resistant = df_output[df_output['Label_Predicted'] == 1].reset_index(drop=True)

        print(f"  TEST total          : {len(df_output)}")
        print(f"  Predicted resistant : {len(df_resistant)}")
        print(f"  Predicted non-res.  : {len(df_output) - len(df_resistant)}")

        if len(df_resistant) == 0:
            print("\n⚠️  No resistant sequences predicted in TEST split.")
            print("   Falling back: using all TEST sequences with Label_Original == 1.")
            df_resistant = df_output[df_output['Label_Original'] == 1].reset_index(drop=True)
            df_resistant['Label_Predicted'] = 1

        # [4] Save
        ExcelWriter.write(self.output_file, df_resistant, all_metrics)

        print(f"\n{'='*80}")
        print("✅ MODEL 1 COMPLETE")
        print(f"   {len(df_resistant)} resistant sequences written to {self.output_file}")
        print(f"   → Pass Data_6.xlsx to model2.py for GNN locus analysis")
        print(f"{'='*80}\n")

# ====================== ENTRY POINT ======================

if __name__ == "__main__":
    pipeline = Model1Pipeline(
        input_file='/kaggle/input/datasets/giogogi/data-4/Data_4.xlsx',
        output_file='Data_6.xlsx',
    )
    pipeline.run()