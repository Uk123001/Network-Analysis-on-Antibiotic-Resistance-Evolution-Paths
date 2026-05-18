"""
Step 1: Data Collection - NCBI optrA Gene Sequences
Retrieves optrA sequences from NCBI and extracts first 500-1000 bp
Output: Data_1.xlsx
"""
# Imported modules:
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from Bio import Entrez
import time
import requests
from bs4 import BeautifulSoup
import hashlib


# Configure Entrez
Entrez.email = "uthamkumar012@gmail.com"  # Required by NCBI
Entrez.api_key = "b4371ae23d0d2b249a43c9f83d803a53cc08"  # Optional but recommended


class NCBIRetriever:
    """Retrieves optrA gene sequences from NCBI"""
    
    def __init__(self, max_records=1500):
        self.max_records = max_records
        self.sequences = []
    
    def search_sequences(self):
        """Search for optrA sequences in NCBI nucleotide database"""
        print("[NCBI] Searching for optrA gene sequences...")
        try:
            handle = Entrez.esearch(db="nucleotide", term="optrA", retmax=self.max_records)
            record = Entrez.read(handle)
            handle.close()
            
            count = int(record["Count"])
            idlist = record["IdList"]
            
            print(f"[NCBI] Total optrA sequences found: {count}")
            print(f"[NCBI] Retrieving: {len(idlist)} sequences")
            
            return idlist
        
        except Exception as e:
            print(f"[NCBI] Search error: {type(e).__name__}: {e}")
            return []
    
    def fetch_sequences(self, idlist):
        """Fetch full sequence records from NCBI"""
        print(f"[NCBI] Fetching {len(idlist)} sequences...")
        
        for i, seq_id in enumerate(idlist):
            try:
                handle = Entrez.efetch(db="nucleotide", id=seq_id, rettype="fasta", retmode="text")
                record = handle.read()
                handle.close()
                
                seq_data = self._parse_fasta(record, seq_id)
                if seq_data:
                    self.sequences.append(seq_data)
                
                if (i + 1) % 50 == 0:
                    print(f"[NCBI] Processed {i + 1}/{len(idlist)}")
                    time.sleep(0.5)
            
            except Exception as e:
                print(f"[NCBI] Error fetching {seq_id}: {e}")
                continue
        
        print(f"[NCBI] Successfully retrieved: {len(self.sequences)} sequences")
        return self.sequences
    
    def _parse_fasta(self, fasta_text, seq_id):
        """Parse FASTA format and extract first 500-1000 bp"""
        try:
            lines = fasta_text.strip().split('\n')
            if not lines:
                return None
            
            header = lines[0].replace('>', f'optrA_NCBI_{seq_id}_')
            full_sequence = ''.join(lines[1:])
            
            # Extract first 500-1000 bp
            if len(full_sequence) < 500:
                return None  # Skip if sequence is too short
            
            sequence = full_sequence[:1000]  # Take first 1000 bp
            
            if sequence:
                return {
                    'header': header,
                    'sequence': sequence.upper(),
                    'sequence_length': len(sequence),
                    'source': 'NCBI'
                }
        
        except Exception as e:
            print(f"[NCBI] Parse error for {seq_id}: {e}")
        
        return None


class GenBankRetriever:
    """Retrieves optrA gene sequences from GenBank"""
    
    def __init__(self, max_records=1500):
        self.max_records = max_records
        self.sequences = []
    
    def search_and_fetch(self):
        """Search GenBank and fetch optrA sequences"""
        print("[GenBank] Searching for optrA gene sequences...")
        
        try:
            handle = Entrez.esearch(db="nucleotide", term="optrA", retmax=self.max_records)
            record = Entrez.read(handle)
            handle.close()
            
            count = int(record["Count"])
            idlist = record["IdList"]
            
            print(f"[GenBank] Total optrA sequences found: {count}")
            print(f"[GenBank] Retrieving: {len(idlist)} sequences")
            
            for i, seq_id in enumerate(idlist):
                try:
                    handle = Entrez.efetch(db="nucleotide", id=seq_id, rettype="fasta", retmode="text")
                    fasta_record = handle.read()
                    handle.close()
                    
                    seq_data = self._parse_fasta(fasta_record, seq_id)
                    if seq_data:
                        self.sequences.append(seq_data)
                    
                    if (i + 1) % 50 == 0:
                        print(f"[GenBank] Processed {i + 1}/{len(idlist)}")
                        time.sleep(0.5)
                
                except Exception as e:
                    print(f"[GenBank] Error fetching {seq_id}: {e}")
                    continue
            
            print(f"[GenBank] Successfully retrieved: {len(self.sequences)} sequences")
            return self.sequences
        
        except Exception as e:
            print(f"[GenBank] Search error: {type(e).__name__}: {e}")
            return []
    
    def _parse_fasta(self, fasta_text, seq_id):
        """Parse FASTA format and extract first 500-1000 bp"""
        try:
            lines = fasta_text.strip().split('\n')
            if not lines:
                return None
            
            header = lines[0].replace('>', f'optrA_GenBank_{seq_id}_')
            full_sequence = ''.join(lines[1:])
            
            # Extract first 500-1000 bp
            if len(full_sequence) < 500:
                return None  # Skip if sequence is too short
            
            sequence = full_sequence[:1000]  # Take first 1000 bp
            
            if sequence:
                return {
                    'header': header,
                    'sequence': sequence.upper(),
                    'sequence_length': len(sequence),
                    'source': 'GenBank'
                }
        
        except Exception as e:
            print(f"[GenBank] Parse error for {seq_id}: {e}")
        
        return None


class DataCollectionPipeline:
    """Main pipeline for data collection"""
    
    def __init__(self, output_file='Data_1.xlsx'):
        self.output_file = output_file
        self.all_sequences = []
    
    def run(self):
        """Execute the data collection pipeline"""
        print("=" * 70)
        print("STEP 1: DATA COLLECTION - optrA Gene Sequences (500-1000 bp)")
        print("=" * 70)
        
        # Retrieve from NCBI
        print("\n[1] Retrieving from NCBI...")
        ncbi_retriever = NCBIRetriever(max_records=1500)
        ncbi_ids = ncbi_retriever.search_sequences()
        ncbi_seqs = ncbi_retriever.fetch_sequences(ncbi_ids)
        self.all_sequences.extend(ncbi_seqs)
        
        # Retrieve from GenBank
        print("\n[2] Retrieving from GenBank...")
        genbank_retriever = GenBankRetriever(max_records=1500)
        genbank_seqs = genbank_retriever.search_and_fetch()
        self.all_sequences.extend(genbank_seqs)
        
        print(f"\n[3] Total sequences collected: {len(self.all_sequences)}")
        
        # Save to Excel
        print("\n[4] Saving to Excel file...")
        self._save_to_excel()
        print(f"    Saved to: {self.output_file}")
        
        print("\n" + "=" * 70)
        print(f"COMPLETE: {len(self.all_sequences)} optrA sequences ready for analysis")
        print("=" * 70)
    
    def _save_to_excel(self):
        """Save sequences to Excel"""
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sequences"
        
        # Headers
        headers = ['Header', 'Sequence', 'Sequence_Length', 'Source']
        sheet.append(headers)
        
        # Format header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for seq_data in self.all_sequences:
            sheet.append([
                seq_data['header'],
                seq_data['sequence'],
                seq_data['sequence_length'],
                seq_data['source']
            ])
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 70
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 12
        
        # Add summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet['A1'] = "optrA Gene Sequences - Collection Summary"
        summary_sheet['A1'].font = Font(bold=True, size=12)
        
        ncbi_count = sum(1 for s in self.all_sequences if s['source'] == 'NCBI')
        genbank_count = sum(1 for s in self.all_sequences if s['source'] == 'GenBank')
        
        summary_sheet['A3'] = "Total Sequences Collected:"
        summary_sheet['B3'] = len(self.all_sequences)
        
        summary_sheet['A4'] = "From NCBI:"
        summary_sheet['B4'] = ncbi_count
        
        summary_sheet['A5'] = "From GenBank:"
        summary_sheet['B5'] = genbank_count
        
        summary_sheet['A6'] = "Sequence Length:"
        summary_sheet['B6'] = "500-1000 bp (first 1000 bp extracted)"
        
        summary_sheet.column_dimensions['A'].width = 35
        summary_sheet.column_dimensions['B'].width = 40
        
        wb.save(self.output_file)


if __name__ == "__main__":
    pipeline = DataCollectionPipeline(output_file='Data_1.xlsx')
    pipeline.run()